"""
ArionComply — Document Injection Pipeline Orchestrator  (Phase 2 — with trace)
Entry point for Path B document intake.

Usage:
  # Process a single file
  python3 rag/intake/doc_pipeline.py --file /path/to/policy.pdf --tenant-id <UUID>

  # Process a directory
  python3 rag/intake/doc_pipeline.py --dir /path/to/docs/ --tenant-id <UUID>

  # Dry run (no DB writes)
  python3 rag/intake/doc_pipeline.py --file policy.pdf --tenant-id <UUID> --dry-run

  # With trace output
  python3 rag/intake/doc_pipeline.py --file policy.pdf --tenant-id <UUID> --trace
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
import uuid
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent.parent / ".env")

from rag.intake.models import ExtractionPath, PipelineResult
from rag.intake.readers import read_document
from rag.intake.enricher import enrich
from rag.intake.extractor import extract
from rag.intake.posture_writer import write_findings, update_upload_status

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".doc", ".xlsx", ".xlsm", ".xls", ".txt", ".csv", ".md"}


# =============================================================================
# TRACE WRITER
# All trace writes are best-effort — a failure here never fails the pipeline.
# =============================================================================

class IntakeTracer:
    """
    Writes one row to intake_trace_log per pipeline stage.
    All rows for one file share the same trace_id.
    Silently no-ops if DB is unavailable or table doesn't exist.
    """

    def __init__(
        self,
        db_url:    str,
        trace_id:  str,
        tenant_id: str,
        upload_id: str,
        filename:  str,
        enabled:   bool = True,
    ):
        self.db_url    = db_url
        self.trace_id  = trace_id
        self.tenant_id = tenant_id
        self.upload_id = upload_id
        self.filename  = filename
        self.enabled   = enabled
        self._t_start  = time.time()

    def write(
        self,
        stage:        str,
        stage_ms:     int,
        status:       str = "ok",
        error_type:   Optional[str] = None,
        error_detail: Optional[str] = None,
        **metrics,
    ) -> None:
        """
        Write one trace row. Keyword args map directly to intake_trace_log columns.
        Never raises — exceptions are logged at WARNING so schema drift (e.g. a
        stage missing from the CHECK constraint) surfaces in api.log.
        """
        if not self.enabled:
            return

        total_ms = int((time.time() - self._t_start) * 1000)

        row = {
            "trace_id":     self.trace_id,
            "tenant_id":    self.tenant_id,
            "upload_id":    self.upload_id,
            "filename":     self.filename,
            "stage":        stage,
            "stage_status": status,
            "stage_ms":     stage_ms,
            "total_ms":     total_ms,
            "error_type":   error_type,
            "error_detail": error_detail[:500] if error_detail else None,
        }
        # Merge stage-specific metrics
        allowed = {
            "token_estimate", "page_count", "section_count",
            "extraction_path", "doc_type", "standard_ids", "explicit_refs_found",
            "llm_calls", "findings_raw", "findings_kept",
            "findings_written", "posture_created", "posture_updated", "posture_skipped",
            # Stage 4.5 (xfw_proposer) metrics
            "proposals_written", "proposals_skipped", "xfw_targets",
            # Stage 3 (extract) quality metrics — schema_v35
            "dropped_low_conf", "dropped_short_quote", "dropped_hallucinated",
            "dropped_unknown_ref",
            "markdown_chars", "paragraph_chars", "candidate_controls",
            # schema_v36 — tight scope from top-confidence doc_mappings match
            "primary_candidate_controls",
            # schema_v37 — doc_mappings match count (0 means fallback to
            # legacy _scope_controls; surface via /admin/intake/unmatched-patterns)
            "doc_mappings_match_count",
            # schema_v41 — doc-shape filter signals
            #   dropped_questionnaire (from 2026-06-12 filter)
            #   skipped_as_toc        (from 2026-06-15 filter)
            "dropped_questionnaire", "skipped_as_toc",
            # schema_v42 — extractor↔catalog crosscheck signals
            "crosscheck_confirmed", "crosscheck_disagreements", "crosscheck_unavailable",
            # schema_v44 — workbook sheet classification (Part A: structured-extraction retirement)
            "workbook_sheets_total", "workbook_sheets_mapped",
            "workbook_sheets_unmapped", "workbook_unmapped_sheets",
            "workbook_skipped_meta_sheets",
            # schema_v48 — extraction yield + pass-2 telemetry (closes
            # the silent-loss gap on the recall pass; see
            # [[llm-narrative-under-discovery-audit-2026-06-26]])
            "distinct_musts_bound", "leaf_musts_in_scope", "yield_ratio_pct",
            "pass2_leaves_targeted", "pass2_findings",
        }
        for k, v in metrics.items():
            if k in allowed:
                row[k] = v

        try:
            import psycopg2
            conn = psycopg2.connect(self.db_url)
            cols = [k for k, v in row.items() if v is not None]
            vals = [row[k] for k in cols]
            placeholders = ", ".join(["%s"] * len(cols))
            col_names    = ", ".join(cols)
            with conn.cursor() as cur:
                cur.execute(
                    f"INSERT INTO intake_trace_log ({col_names}) VALUES ({placeholders})",
                    vals,
                )
            conn.commit()
            conn.close()
            logger.debug(f"  [trace] {stage} {status} {stage_ms}ms")
        except Exception as e:
            logger.warning(f"  [trace] write failed (non-fatal): {type(e).__name__}: {e}")


# =============================================================================
# MAIN PIPELINE
# =============================================================================

class DocumentPipeline:
    """
    Orchestrates all stages of the document injection pipeline.

    Stages:
      1. Read    — extract text and structure from file
      2. Enrich  — classify, detect standard, scan for refs, decide path
      3. Extract — LLM extraction (or structured parse for XLSX/CSV)
      4. Write   — document_findings + posture_controls aggregation
    """

    def __init__(
        self,
        db_url:   str,
        api_key:  str,
        dry_run:  bool = False,
        verbose:  bool = False,
        trace:    bool = False,
    ):
        self.db_url  = db_url
        self.api_key = api_key
        self.dry_run = dry_run
        self.trace   = trace

        level = logging.DEBUG if verbose else logging.INFO
        logging.basicConfig(
            format  = "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
            datefmt = "%H:%M:%S",
            level   = level,
        )

        self._controls_cache: dict[str, list[dict]] = {}

    def run(
        self,
        file_path:            str,
        tenant_id:            str,
        upload_id:            Optional[str]        = None,
        original_filename:    Optional[str]        = None,
        user_id:              Optional[str]        = None,
        declared_standard_ids: Optional[list[str]] = None,
        declared_evidence_type: Optional[str]      = None,
    ) -> PipelineResult:
        """
        Process one document.
        Returns PipelineResult with status and counts.

        original_filename: the user-facing name (e.g. "Access_Control_Policy.docx").
        Used for registry lookup; falls back to the disk basename. API uploads
        store the file with a UUID name, so without this the DOC-prefix /
        title matchers can never link to the pre-registered client_documents row.
        """
        t_start    = time.time()
        file_path  = str(Path(file_path).resolve())
        file_name  = original_filename or Path(file_path).name
        trace_id   = str(uuid.uuid4())
        upload_id  = upload_id or str(uuid.uuid4())

        # ── Tracer — always created, enabled flag controls writes ─────────────
        tracer = IntakeTracer(
            db_url    = self.db_url,
            trace_id  = trace_id,
            tenant_id = tenant_id,
            upload_id = upload_id,
            filename  = file_name,
            enabled   = self.trace and not self.dry_run,
        )

        if not Path(file_path).exists():
            tracer.write("failed", 0, status="error",
                         error_type="FileNotFound",
                         error_detail=f"File not found: {file_path}")
            return PipelineResult(
                upload_id       = upload_id,
                tenant_id       = tenant_id,
                document_name   = file_name,
                doc_type        = None,
                standard_ids    = [],
                extraction_path = "failed",
                findings_count  = 0,
                controls_assessed = [],
                controls_updated  = [],
                status = "failed",
                error  = f"File not found: {file_path}",
            )

        logger.info(f"{'[DRY RUN] ' if self.dry_run else ''}Processing: {file_name}"
                    f"  trace_id={trace_id[:8]}")

        try:
            # ── Stage 1: Read ─────────────────────────────────────────────────
            logger.info(f"Stage 1: Reading {file_name}")
            t1 = time.time()
            doc = read_document(
                file_path,
                upload_id         = upload_id,
                original_filename = file_name,
            )
            s1_ms = int((time.time() - t1) * 1000)

            tracer.write(
                "read", s1_ms,
                token_estimate = doc.token_estimate,
                page_count     = doc.page_count,
                section_count  = len(doc.raw_sections),
            )

            # ── Stage 1.5: markdown-content dedup ─────────────────────────────
            # Layer 2 of the v19 idempotency contract. Layer 1 (source bytes)
            # ran at the API edge. If two different exports of the same content
            # slip through (e.g. DOCX vs PDF, same paragraphs), the markdown
            # hash catches it here — before any LLM cost is incurred.
            if doc.markdown and not self.dry_run:
                import hashlib as _hl_dup
                _md_sha_pre = _hl_dup.sha256(doc.markdown.encode("utf-8")).hexdigest()
                _canonical = self._find_markdown_duplicate(
                    tenant_id, _md_sha_pre, exclude_upload_id=upload_id,
                )
                if _canonical:
                    msg = (
                        f"Content already present (matches upload "
                        f"{_canonical[:8]}). No findings written."
                    )
                    logger.info(
                        f"Pipeline aborted as duplicate (markdown match): "
                        f"upload_id={upload_id[:8]} dup_of={_canonical[:8]}"
                    )
                    self._mark_duplicate(upload_id, _canonical, file_path)
                    tracer.write(
                        "duplicate", 0, status="duplicate",
                        error_type="markdown_dup",
                        error_detail=msg,
                        dup_of_upload_id=_canonical,
                    )
                    return PipelineResult(
                        upload_id       = upload_id,
                        tenant_id       = tenant_id,
                        document_name   = file_name,
                        doc_type        = doc.doc_type,
                        standard_ids    = doc.standard_ids,
                        extraction_path = "duplicate",
                        findings_count  = 0,
                        controls_assessed = [],
                        controls_updated  = [],
                        status      = "duplicate",
                        error       = msg,
                        duration_ms = int((time.time() - t_start) * 1000),
                    )

            # ── Stage 2: Enrich ───────────────────────────────────────────────
            logger.info(f"Stage 2: Enriching — ~{doc.token_estimate:,} tokens")
            if declared_standard_ids or declared_evidence_type:
                logger.info(
                    f"Stage 2: tenant hints — standards={declared_standard_ids} "
                    f"type={declared_evidence_type!r}"
                )
            t2 = time.time()
            doc = enrich(
                doc,
                api_key            = self.api_key,
                hint_standard_ids  = declared_standard_ids,
                hint_evidence_type = declared_evidence_type,
            )
            s2_ms = int((time.time() - t2) * 1000)

            tracer.write(
                "enrich", s2_ms,
                extraction_path     = doc.extraction_path.value,
                doc_type            = doc.doc_type,
                standard_ids        = doc.standard_ids,
                explicit_refs_found = len(doc.explicit_refs),
            )

            # ── Manual review branch ──────────────────────────────────────────
            if doc.extraction_path == ExtractionPath.MANUAL_REVIEW:
                msg = (
                    f"Document too large ({doc.token_estimate:,} tokens). "
                    f"Split into individual policies before re-uploading."
                )
                logger.warning(f"Manual review required: {file_name} — {msg}")
                tracer.write("failed", 0, status="manual_review",
                             error_type="DocumentTooLarge", error_detail=msg)
                if not self.dry_run:
                    self._update_status(upload_id, "manual_review", 0, msg)
                return PipelineResult(
                    upload_id       = upload_id,
                    tenant_id       = tenant_id,
                    document_name   = file_name,
                    doc_type        = doc.doc_type,
                    standard_ids    = doc.standard_ids,
                    extraction_path = "manual",
                    findings_count  = 0,
                    controls_assessed = [],
                    controls_updated  = [],
                    status      = "manual_review",
                    error       = msg,
                    duration_ms = int((time.time() - t_start) * 1000),
                )

            # ── Stage 3: Get controls + Extract ──────────────────────────────
            # Ship 40'.b — under consensus extraction, widen standard_ids to
            # all Neo4j-loaded standards so cross-framework controls
            # (e.g. GDPR obligations paired to ISO 27701 extensions) are
            # available for consensus signals. Enricher assigns per-doc
            # standard_ids based on doc content; consensus needs the full
            # framework surface to attribute cross-framework evidence.
            std_ids_for_load = doc.standard_ids
            if os.getenv("USE_CONSENSUS_EXTRACTION") == "1":
                std_ids_for_load = self._all_graph_standards() or doc.standard_ids
                if std_ids_for_load != doc.standard_ids:
                    logger.info(
                        f"Consensus scope widening: {doc.standard_ids} → "
                        f"{std_ids_for_load}"
                    )
            logger.info(f"Stage 3: Loading controls for {std_ids_for_load}")
            controls = self._get_controls(std_ids_for_load)
            logger.info(
                f"Stage 3: Extracting via {doc.extraction_path.value} path "
                f"({len(controls)} controls available)"
            )
            if not self.dry_run:
                self._update_status(upload_id, "processing", 0)

            # ── Templated-xlsx tenant cross-check ─────────────────────
            # If the reader detected our template (via _arion_meta or
            # filename-fallback), validate the embedded tenant_id (when
            # present) matches the uploading tenant. Mismatch → strip
            # the meta + log warning so the file falls through to the
            # generic workbook lane.
            #
            # Filename-fallback path has tenant_id='' (no source of
            # truth for it once _arion_meta is gone); auth + RLS still
            # scope the upload to the uploading tenant correctly.
            tx_meta = doc.extraction_metrics.get("templated_xlsx_meta")
            if tx_meta:
                meta_tenant = tx_meta.get("tenant_id") or ""
                if meta_tenant and meta_tenant != tenant_id:
                    logger.warning(
                        f"templated_xlsx tenant mismatch on {file_name}: "
                        f"meta tenant_id={meta_tenant!r} vs uploading "
                        f"tenant_id={tenant_id!r}. Stripping meta — file will "
                        f"be processed via generic workbook lane."
                    )
                    doc.extraction_metrics.pop("templated_xlsx_meta", None)
                elif tx_meta.get("source") == "filename_fallback":
                    logger.info(
                        f"templated_xlsx filename-fallback on {file_name}: "
                        f"_arion_meta missing; resolved leaf_id="
                        f"{tx_meta.get('leaf_id')!r} via filename convention."
                    )

            t3 = time.time()
            findings = extract(doc, controls, self.api_key)
            s3_ms = int((time.time() - t3) * 1000)

            # Prefer the extractor's tracked LLM-call count over the
            # section-count estimate — section_based now sometimes
            # rebuilds sections from markdown chunks, so len(raw_sections)
            # under-counts vs the actual call count.
            llm_calls = doc.extraction_metrics.get("llm_calls")
            if llm_calls is None:
                llm_calls = 1 if doc.extraction_path != ExtractionPath.SECTION_BASED \
                              else max(1, len(doc.raw_sections))

            tracer.write(
                "extract", s3_ms,
                llm_calls    = llm_calls,
                findings_raw = len(findings),
                findings_kept= len(findings),
                # Quality telemetry — schema_v35
                dropped_low_conf     = doc.extraction_metrics.get("dropped_low_conf"),
                dropped_short_quote  = doc.extraction_metrics.get("dropped_short_quote"),
                dropped_hallucinated = doc.extraction_metrics.get("dropped_hallucinated"),
                dropped_unknown_ref  = doc.extraction_metrics.get("dropped_unknown_ref"),
                markdown_chars       = doc.extraction_metrics.get("markdown_chars"),
                paragraph_chars      = doc.extraction_metrics.get("paragraph_chars"),
                candidate_controls   = doc.extraction_metrics.get("candidate_controls"),
                primary_candidate_controls = doc.extraction_metrics.get("primary_candidate_controls"),
                doc_mappings_match_count = doc.extraction_metrics.get("doc_mappings_match_count"),
                # schema_v41 — doc-shape filter signals
                dropped_questionnaire = doc.extraction_metrics.get("dropped_questionnaire"),
                skipped_as_toc        = doc.extraction_metrics.get("skipped_as_toc"),
                # schema_v42 — extractor↔catalog crosscheck signals
                crosscheck_confirmed     = doc.extraction_metrics.get("crosscheck_confirmed"),
                crosscheck_disagreements = doc.extraction_metrics.get("crosscheck_disagreements"),
                crosscheck_unavailable   = doc.extraction_metrics.get("crosscheck_unavailable"),
                # schema_v44 — workbook sheet classification (xlsx/xlsm only)
                workbook_sheets_total        = doc.extraction_metrics.get("workbook_sheets_total"),
                workbook_sheets_mapped       = doc.extraction_metrics.get("workbook_sheets_mapped"),
                workbook_sheets_unmapped     = doc.extraction_metrics.get("workbook_sheets_unmapped"),
                workbook_unmapped_sheets     = doc.extraction_metrics.get("workbook_unmapped_sheets"),
                workbook_skipped_meta_sheets = doc.extraction_metrics.get("workbook_skipped_meta_sheets"),
                # schema_v48 — extraction yield + pass-2 telemetry. Closes
                # the silent-loss gap on the recall pass; see
                # [[llm-narrative-under-discovery-audit-2026-06-26]].
                distinct_musts_bound  = doc.extraction_metrics.get("distinct_musts_bound"),
                leaf_musts_in_scope   = doc.extraction_metrics.get("leaf_musts_in_scope"),
                yield_ratio_pct       = doc.extraction_metrics.get("yield_ratio_pct"),
                pass2_leaves_targeted = doc.extraction_metrics.get("pass2_leaves_targeted"),
                pass2_findings        = doc.extraction_metrics.get("pass2_findings"),
            )

            logger.info(f"Extracted {len(findings)} findings from {file_name}")

            if self.dry_run:
                self._print_dry_run(findings, doc)
                return PipelineResult(
                    upload_id       = upload_id,
                    tenant_id       = tenant_id,
                    document_name   = file_name,
                    doc_type        = doc.doc_type,
                    standard_ids    = doc.standard_ids,
                    extraction_path = doc.extraction_path.value,
                    findings_count  = len(findings),
                    controls_assessed = list({f.control_ref for f in findings}),
                    controls_updated  = [],
                    status      = "dry_run",
                    duration_ms = int((time.time() - t_start) * 1000),
                )

            # ── Stage 4: Write ────────────────────────────────────────────────
            logger.info(f"Stage 4: Writing {len(findings)} findings to DB")
            t4 = time.time()

            # Bundle file/content metadata so the writer can stamp it on
            # client_documents alongside the registry linkage update.
            import hashlib
            import mimetypes
            _path_obj = Path(file_path)
            try:
                _file_bytes = _path_obj.read_bytes()
                _file_size  = len(_file_bytes)
                _sha256     = hashlib.sha256(_file_bytes).hexdigest()
            except Exception:
                _file_size, _sha256 = None, None
            _mime, _ = mimetypes.guess_type(file_name)

            # control_refs cached on client_documents store fully-qualified
            # STANDARD:VERSION:REF entries so cross-framework attribution
            # survives without the loader having to assume a framework.
            # The live read in load_uploaded_documents prefers
            # document_findings, but if those are absent the cached column
            # is now framework-correct by itself.
            doc_metadata = {
                "file_size_bytes": _file_size,
                "mime_type":       _mime,
                "checksum_sha256": _sha256,
                "page_count":      doc.page_count,
                "evidence_type":   doc.doc_type,
                "control_refs":    sorted({
                    f"{f.standard_id}:{f.control_ref}"
                    for f in findings
                    if f.control_ref and f.standard_id
                }),
            }

            import psycopg2
            conn = psycopg2.connect(self.db_url)
            try:
                # RLS needs app.tenant_id on this connection. Other tables
                # (document_findings, client_documents, posture_controls)
                # are scoped permissively for arioncomply_app, but
                # tabular_evidence_rows (schema_v47) and tenant_profile
                # (schema_v49) enforce strict RLS so any future read
                # paths can't leak across tenants — INSERT also requires
                # the GUC to be set.
                with conn.cursor() as _cur:
                    _cur.execute("SET app.tenant_id = %s", (tenant_id,))
                # Signal-fusion Wave 3 (2026-07-09): four independent
                # corroboration signals for fingerprint_match auto-approval,
                # each derived from a DIFFERENT view of the doc so their
                # agreement is genuinely informative:
                #   target_controls   — filename + topic tokens (doc_mappings)
                #   semantic_controls — body-content semantic match (musts_arioncomply)
                #   explicit_refs     — doc self-citations (regex-extracted refs)
                #   llm_extract       — LLM ran and produced findings for this control
                # The writer's gate requires ≥2 available signals to agree.
                _tgt_leaves = doc.extraction_metrics.get("target_leaves") or []
                _target_controls: Optional[set[str]] = {
                    lf.get("control_ref") for lf in _tgt_leaves
                    if lf.get("control_ref")
                } or None

                try:
                    from rag.intake.must_embedding_lookup import (
                        semantic_controls_in_scope,
                    )
                    _semantic_controls: Optional[set[str]] = semantic_controls_in_scope(
                        doc_text    = doc.markdown or doc.full_text,
                        tenant_stds = doc.standard_ids or None,
                    ) or None
                except Exception:
                    _semantic_controls = None

                # Explicit refs are stored on ParsedDocument (set during
                # enrichment via extract_refs_from_text). Convert to set of
                # bare control_refs. None means the regex scan ran but found
                # nothing; still a valid signal (0 agreement, 1 available).
                _explicit_refs: Optional[set[str]] = (
                    set(doc.explicit_refs) if doc.explicit_refs is not None else None
                )

                # LLM signal — set of control_refs where the LLM produced any
                # 'extracted' finding in this batch. LLM path leaves
                # inference_source None on the finding object (relies on DB
                # default 'extracted'); templated / fingerprint / workbook /
                # leaf_scan set it explicitly. So "LLM finding" = the object
                # doesn't have any of the explicit deterministic sources set.
                _DETERMINISTIC_SOURCES = {
                    "templated", "fingerprint_match", "workbook",
                    "leaf_scan", "form", "xfw_bridge",
                }
                _llm_extracted: Optional[set[str]] = {
                    f.control_ref for f in findings
                    if getattr(f, "inference_source", None) not in _DETERMINISTIC_SOURCES
                    and f.control_ref
                } or None

                summary = write_findings(
                    findings, tenant_id, upload_id, conn,
                    metadata           = doc_metadata,
                    uploaded_by        = user_id,
                    tabular_rows       = doc.tabular_rows or None,
                    target_controls    = _target_controls,
                    semantic_controls  = _semantic_controls,
                    explicit_refs      = _explicit_refs,
                    llm_extracted      = _llm_extracted,
                )

                # Persist the parsed markdown alongside the findings so the
                # extractor's input is reproducible without the original
                # binary. Skipped silently for formats with no markdown
                # renderer yet (pdf/xlsx/csv/txt).
                if doc.markdown:
                    import hashlib as _hl
                    _md_bytes = doc.markdown.encode("utf-8")
                    _md_sha   = _hl.sha256(_md_bytes).hexdigest()
                    with conn.cursor() as _cur:
                        _cur.execute(
                            """
                            INSERT INTO document_text (
                                upload_id, tenant_id, markdown,
                                markdown_sha256, source_sha256,
                                converter, byte_count
                            ) VALUES (
                                %s::uuid, %s::uuid, %s, %s, %s, %s, %s
                            )
                            ON CONFLICT (upload_id) DO UPDATE SET
                                markdown        = EXCLUDED.markdown,
                                markdown_sha256 = EXCLUDED.markdown_sha256,
                                source_sha256   = EXCLUDED.source_sha256,
                                converter       = EXCLUDED.converter,
                                byte_count      = EXCLUDED.byte_count,
                                parsed_at       = now()
                            """,
                            (
                                upload_id,
                                tenant_id,
                                doc.markdown,
                                _md_sha,
                                doc.source_sha256 or _sha256 or "",
                                doc.converter or "unknown",
                                len(_md_bytes),
                            ),
                        )

                conn.commit()
            except Exception as e:
                conn.rollback()
                raise
            finally:
                try:
                    update_upload_status(
                        upload_id      = upload_id,
                        status         = "completed",
                        findings_count = len(findings),
                        conn           = conn,
                    )
                    conn.commit()
                except Exception:
                    pass
                conn.close()

            # ── Stage 4.5: xfw proposer ───────────────────────────────────────
            # Walk Neo4j IMPLEMENTS edges from each just-written finding and
            # propose mirror findings on xfw-bridged standards (filtered by
            # tenant_evaluation_scope). Proposals land in document_findings with
            # confirmed_by IS NULL — the HITL queue. Failures here are logged
            # and swallowed: Stage 4 has already committed; an xfw failure must
            # not poison the upload.
            t4_5 = time.time()
            _xfw_written = 0
            _xfw_skipped = 0
            _xfw_targets: list[str] = []
            _xfw_error: Optional[str] = None
            _xfw_doc_id  = summary.get("doc_id")
            if not findings or not _xfw_doc_id:
                # Nothing extracted → nothing to walk. Skip silently — this is
                # the empty-extraction path (e.g. LLM failure) and emitting an
                # error here would mask the real Stage 3 problem.
                logger.debug(
                    "Stage 4.5 skipped: no findings or doc_id unresolved "
                    f"(findings={len(findings)}, doc_id={_xfw_doc_id})"
                )
            else:
                try:
                    from rag.intake.xfw_proposer import propose_for_findings
                    from neo4j import GraphDatabase
                    _neo_driver = GraphDatabase.driver(
                        os.getenv("NEO4J_URI", "bolt://localhost:7687"),
                        auth=(os.getenv("NEO4J_USER", "neo4j"),
                              os.getenv("NEO4J_PASSWORD", "")),
                    )
                    _xfw_conn = psycopg2.connect(self.db_url)
                    try:
                        with _xfw_conn.cursor() as _cur:
                            _cur.execute("SET app.tenant_id = %s", (tenant_id,))
                        _xfw_summary = propose_for_findings(
                            tenant_id   = tenant_id,
                            document_id = _xfw_doc_id,
                            findings    = findings,
                            conn        = _xfw_conn,
                            driver      = _neo_driver,
                        )
                        _xfw_conn.commit()
                        _xfw_written = _xfw_summary.proposals_written
                        _xfw_skipped = _xfw_summary.proposals_skipped
                        _xfw_targets = sorted(_xfw_summary.standards_targeted)
                        logger.info(f"Stage 4.5: {_xfw_summary}")
                    except Exception:
                        _xfw_conn.rollback()
                        raise
                    finally:
                        _xfw_conn.close()
                        _neo_driver.close()
                except Exception as e:
                    _xfw_error = f"{type(e).__name__}: {e}"
                    logger.warning(
                        f"xfw_proposer hook failed (Stage 4 already committed): "
                        f"{_xfw_error}"
                    )

            tracer.write(
                "xfw",
                int((time.time() - t4_5) * 1000),
                status            = "error" if _xfw_error else "ok",
                error_detail      = _xfw_error,
                proposals_written = _xfw_written,
                proposals_skipped = _xfw_skipped,
                xfw_targets       = _xfw_targets,
            )

            # ── Stage 4.6: workbook discovery (xlsx/xlsm only) ───────────────
            # Complements the doc-extractor path. The extractor reads xlsx as
            # tokenised text and runs LLM extraction; workbook discovery reads
            # the same file as STRUCTURED rows and produces per-row evidence
            # bound to specific checklist_item_ids. Both paths are useful and
            # not mutually exclusive — discovery adds the structured layer
            # without disturbing the LLM findings already written.
            #
            # Best-effort: any failure here is logged and swallowed. The
            # upload itself succeeded via Stage 4; workbook discovery is a
            # bonus pass.
            t4_6 = time.time()
            _wbd_proposals = 0
            _wbd_findings  = 0
            _wbd_error: Optional[str] = None
            _is_workbook = file_name.lower().endswith((".xlsx", ".xlsm"))
            _wbd_doc_id  = summary.get("doc_id")
            # When extract returns 0 findings (new Part-A behaviour: xlsx/xlsm
            # skip _extract_structured), summary lacks doc_id but the workbook
            # still has a client_documents row from the upload. Look it up by
            # sha256 — safer than filename because dedup normalises bytes.
            if _is_workbook and not _wbd_doc_id and not self.dry_run:
                try:
                    import hashlib
                    with open(file_path, "rb") as _f:
                        _sha = hashlib.sha256(_f.read()).hexdigest()
                    _wbd_conn_lookup = psycopg2.connect(self.db_url)
                    try:
                        with _wbd_conn_lookup.cursor() as _cur:
                            _cur.execute("SET app.tenant_id = %s", (tenant_id,))
                            _cur.execute(
                                """
                                SELECT id::text FROM client_documents
                                WHERE tenant_id = %s::uuid
                                  AND checksum_sha256 = %s
                                  AND is_active = TRUE
                                ORDER BY uploaded_at DESC LIMIT 1
                                """,
                                (tenant_id, _sha),
                            )
                            _row = _cur.fetchone()
                            if _row:
                                _wbd_doc_id = _row[0]
                                logger.info(
                                    f"Stage 4.6: doc_id from sha256 lookup: {_wbd_doc_id}"
                                )
                    finally:
                        _wbd_conn_lookup.close()
                except Exception as _e:
                    logger.warning(f"Stage 4.6: doc_id lookup failed: {_e}")
            if _is_workbook and _wbd_doc_id and not self.dry_run:
                try:
                    import openpyxl
                    from rag.intake.workbook_discovery import discover_workbook
                    from rag.intake.workbook_persistence import persist_proposals
                    from uuid import UUID

                    _wb = openpyxl.load_workbook(
                        file_path, keep_vba=True, data_only=True, read_only=True,
                    )
                    _rows_per_sheet: dict[str, list[list]] = {}
                    for _sheet_name in _wb.sheetnames:
                        _ws = _wb[_sheet_name]
                        _rows_per_sheet[_sheet_name] = [
                            list(r) for r in _ws.iter_rows(values_only=True)
                        ]
                    _wb.close()

                    _proposals = discover_workbook(_rows_per_sheet)

                    if _proposals:
                        _wbd_conn = psycopg2.connect(self.db_url)
                        try:
                            _, _wbd_findings = persist_proposals(
                                _wbd_conn,
                                UUID(tenant_id),
                                file_path,
                                UUID(_wbd_doc_id),
                                _proposals,
                            )
                            _wbd_proposals = len(_proposals)
                            logger.info(
                                f"Stage 4.6: workbook discovery wrote "
                                f"{_wbd_proposals} proposals + {_wbd_findings} "
                                f"findings"
                            )
                        finally:
                            _wbd_conn.close()
                    else:
                        logger.info(
                            "Stage 4.6: workbook discovery matched 0 sheets — "
                            "no proposals to persist"
                        )
                except Exception as e:
                    _wbd_error = f"{type(e).__name__}: {e}"
                    logger.warning(
                        f"workbook_discovery hook failed (Stage 4 already "
                        f"committed): {_wbd_error}"
                    )

            # Trace only when the workbook path actually ran. The
            # tracer's stage column has a CHECK constraint
            # (intake_trace_log_stage_check) that we haven't yet
            # extended to include "workbook_discovery", so writing
            # this stage row on every upload would CheckViolation
            # for non-xlsx files. Until schema_v3X adds the new
            # stage value, gate on _is_workbook. PDF/docx uploads
            # don't need a workbook_discovery telemetry row anyway
            # — the stage genuinely didn't run for them.
            if _is_workbook and not self.dry_run:
                tracer.write(
                    "workbook_discovery",
                    int((time.time() - t4_6) * 1000),
                    status            = "error" if _wbd_error else "ok",
                    error_detail      = _wbd_error,
                    proposals_written = _wbd_proposals,
                    findings_written  = _wbd_findings,
                )

            s4_ms = int((time.time() - t4) * 1000)

            tracer.write(
                "write", s4_ms,
                findings_written = summary.get("written", 0),
                posture_created  = summary.get("posture_created", 0),
                posture_updated  = summary.get("posture_updated", 0),
                posture_skipped  = summary.get("posture_skipped", 0),
            )

            duration_ms = int((time.time() - t_start) * 1000)

            # ── Complete trace row ────────────────────────────────────────────
            tracer.write("complete", 0)

            logger.info(
                f"Complete: {file_name} | "
                f"{len(findings)} findings | "
                f"{summary['posture_updated']} updated | "
                f"{summary['posture_created']} created | "
                f"{summary.get('posture_skipped', 0)} skipped | "
                f"{duration_ms}ms | trace={trace_id[:8]}"
            )

            return PipelineResult(
                upload_id       = upload_id,
                tenant_id       = tenant_id,
                document_name   = file_name,
                doc_type        = doc.doc_type,
                standard_ids    = doc.standard_ids,
                extraction_path = doc.extraction_path.value,
                findings_count  = len(findings),
                controls_assessed = list({f.control_ref for f in findings}),
                controls_updated  = summary.get("controls_assessed", []),
                status      = "extracted",
                duration_ms = duration_ms,
            )

        except Exception as e:
            duration_ms = int((time.time() - t_start) * 1000)
            logger.error(f"Pipeline failed for {file_name}: {e}", exc_info=True)
            tracer.write(
                "failed", duration_ms, status="error",
                error_type   = type(e).__name__,
                error_detail = str(e),
            )
            if not self.dry_run:
                try:
                    self._update_status(upload_id, "failed", 0, str(e))
                except Exception:
                    pass
                # Ship 3'.e producer: notify tenant of the upload failure.
                # Dedup key is (kind, related_entity_id=upload_id) via the
                # partial unique index, so retries on the same upload_id
                # collapse. Severity 'medium' by default; caller can escalate
                # to 'high' later if repeat-failure detection is added.
                try:
                    from rag.cascade.notify import notify as _notify
                    import psycopg2 as _pg
                    _conn = _pg.connect(self.db_url)
                    try:
                        with _conn.cursor() as _cur:
                            _cur.execute(
                                "SELECT set_config('app.tenant_id', %s, TRUE)",
                                (tenant_id,),
                            )
                            _notify(
                                _cur,
                                tenant_id           = tenant_id,
                                kind                = "upload_failed",
                                title               = f"Upload failed: {file_name}",
                                body                = (
                                    f"{type(e).__name__}: {str(e)[:200]}"
                                ),
                                severity            = "medium",
                                related_entity_kind = "document_upload",
                                related_entity_id   = upload_id,
                                related_control_ref = None,
                                related_event_type  = "upload_failed",
                            )
                        _conn.commit()
                    finally:
                        _conn.close()
                except Exception as _e:
                    logger.warning(
                        f"upload_failed notify failed: {type(_e).__name__}: {_e}"
                    )
            return PipelineResult(
                upload_id       = upload_id,
                tenant_id       = tenant_id,
                document_name   = file_name,
                doc_type        = None,
                standard_ids    = [],
                extraction_path = "failed",
                findings_count  = 0,
                controls_assessed = [],
                controls_updated  = [],
                status      = "failed",
                error       = str(e),
                duration_ms = duration_ms,
            )

    def run_directory(
        self,
        directory: str,
        tenant_id: str,
    ) -> list[PipelineResult]:
        """Process all supported files in a directory."""
        results = []
        files   = sorted(
            p for p in Path(directory).iterdir()
            if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS
        )

        if not files:
            logger.warning(f"No supported files found in {directory}")
            return []

        logger.info(f"Processing {len(files)} files from {directory}")
        for i, file_path in enumerate(files, 1):
            logger.info(f"[{i}/{len(files)}] {file_path.name}")
            result = self.run(str(file_path), tenant_id)
            results.append(result)
            self._print_result(result)

        self._print_summary(results)
        return results

    # ── Helpers ───────────────────────────────────────────────────────────────

    _graph_standards_cache: Optional[list[str]] = None

    def _all_graph_standards(self) -> list[str]:
        """Return all standard_ids present in Neo4j RequirementNode.

        Cached across calls in this instance. Silent fallback to []
        on Neo4j error — caller degrades to per-doc standard_ids."""
        if self._graph_standards_cache is not None:
            return self._graph_standards_cache
        try:
            from neo4j import GraphDatabase
            driver = GraphDatabase.driver(
                os.getenv("NEO4J_URI"),
                auth=(os.getenv("NEO4J_USER"), os.getenv("NEO4J_PASSWORD")),
            )
            try:
                with driver.session() as s:
                    result = s.run(
                        "MATCH (n:RequirementNode) "
                        "RETURN DISTINCT n.standard_id AS sid ORDER BY sid"
                    )
                    self._graph_standards_cache = [r["sid"] for r in result if r["sid"]]
            finally:
                driver.close()
        except Exception as e:
            logger.warning(
                f"_all_graph_standards Neo4j query failed: "
                f"{type(e).__name__}: {e}"
            )
            self._graph_standards_cache = []
        return self._graph_standards_cache

    def _get_controls(self, standard_ids: list[str]) -> list[dict]:
        all_controls = []
        for std in standard_ids:
            if std in self._controls_cache:
                all_controls.extend(self._controls_cache[std])
                continue
            try:
                controls = self._load_controls_from_neo4j(std)
                self._controls_cache[std] = controls
                all_controls.extend(controls)
                logger.info(f"Loaded {len(controls)} controls for {std}")
            except Exception as e:
                logger.warning(f"Could not load controls for {std}: {e}")

        # Phase 3 (framework role model, 2026-07-05): exclude OBLIGATION
        # controls that a PROGRAM/EXTENSION in scope already demonstrates.
        # Those get propagated deterministically via DEMONSTRATES in
        # posture_loader._apply_demonstrates_overlay; letting the LLM
        # also extract them directly reintroduces the pre-Phase 3
        # multi-framework guessing bias. Obligations WITHOUT a
        # DEMONSTRATES source in tenant scope stay in the list so
        # direct extraction still works for pure-legal content
        # (Art.7 consent mechanics, Art.30 records, etc).
        #
        # Ship 40'.a — bypass Phase 3 filter under consensus extraction.
        # Phase 3's "prevent LLM guessing bias" rationale applied to the
        # OLD LLM discovery pipeline. Consensus has its own discipline:
        # fingerprint excerpt requirement + 8-signal aggregator + bounded
        # LLM arbiter. Direct cross-framework findings become an auditor-
        # facing feature. DEMONSTRATES overlay in posture_loader remains
        # as belt-and-suspenders backstop when consensus doesn't surface
        # a direct finding.
        if os.getenv("USE_CONSENSUS_EXTRACTION") == "1":
            logger.info(
                f"Phase 3 filter BYPASSED under consensus "
                f"({len(all_controls)} controls in scope)"
            )
            return all_controls

        filtered = self._filter_demonstrated_obligations(all_controls, standard_ids)
        excluded = len(all_controls) - len(filtered)
        if excluded:
            logger.info(
                f"Phase 3 filter: {excluded} obligation controls excluded "
                f"(demonstrated by PROGRAM/EXTENSION in scope; will be "
                f"propagated via DEMONSTRATES at posture-load time)"
            )
        return filtered

    def _filter_demonstrated_obligations(
        self, controls: list[dict], standard_ids: list[str],
    ) -> list[dict]:
        """Return controls minus OBLIGATION rows whose DEMONSTRATES source
        is in the tenant's PROGRAM/EXTENSION scope.

        Silent fallback: any Neo4j failure returns the input list
        unchanged — Phase 3 is an overlay, not a hard dependency."""
        try:
            from neo4j import GraphDatabase
            driver = GraphDatabase.driver(
                os.getenv("NEO4J_URI"),
                auth=(os.getenv("NEO4J_USER"), os.getenv("NEO4J_PASSWORD")),
            )
            try:
                with driver.session() as s:
                    result = s.run(
                        """
                        MATCH (src:RequirementNode)-[:DEMONSTRATES]->(tgt:RequirementNode)
                        WHERE src.standard_id IN $stds
                          AND tgt.role_owner = 'obligation'
                        RETURN DISTINCT tgt.ref AS ref, tgt.standard_id AS std
                        """,
                        stds=standard_ids,
                    )
                    excluded_keys = {(r["std"], r["ref"]) for r in result}
            finally:
                driver.close()
        except Exception as e:
            logger.warning(
                f"Phase 3 obligation filter skipped ({type(e).__name__}: {e})"
            )
            return controls

        if not excluded_keys:
            return controls
        return [
            c for c in controls
            if (c.get("standard_id"), c.get("ref")) not in excluded_keys
        ]

    def _load_controls_from_neo4j(self, standard_id: str) -> list[dict]:
        neo4j_uri  = os.getenv("NEO4J_URI",      "bolt://localhost:7687")
        neo4j_user = os.getenv("NEO4J_USER",     "neo4j")
        neo4j_pass = os.getenv("NEO4J_PASSWORD", "")
        try:
            from neo4j import GraphDatabase
            driver = GraphDatabase.driver(neo4j_uri, auth=(neo4j_user, neo4j_pass))
            with driver.session() as s:
                result = s.run("""
                    MATCH (n:RequirementNode)
                    WHERE n.standard_id = $std
                    RETURN n.ref AS ref,
                           n.title AS title,
                           n.standard_id AS standard_id
                    ORDER BY n.ref
                """, std=standard_id)
                controls = [
                    {
                        "ref":         row["ref"],
                        "title":       row["title"] or row["ref"],
                        "standard_id": row["standard_id"],
                    }
                    for row in result if row["ref"]
                ]
            driver.close()
            return controls
        except Exception as e:
            logger.warning(f"Neo4j unavailable: {e} — using empty control list")
            return []

    def _update_status(
        self,
        upload_id: str,
        status:    str,
        count:     int,
        error:     Optional[str] = None,
    ) -> None:
        try:
            import psycopg2
            conn = psycopg2.connect(self.db_url)
            update_upload_status(upload_id, status, count, conn, error)
            conn.commit()
            conn.close()
        except Exception as e:
            logger.debug(f"Status update failed: {e}")

    def _find_markdown_duplicate(
        self,
        tenant_id:           str,
        markdown_sha256:     str,
        exclude_upload_id:   str,
    ) -> Optional[str]:
        """Return canonical upload_id if another non-duplicate row already
        holds this markdown hash for the tenant, else None."""
        try:
            import psycopg2
            conn = psycopg2.connect(self.db_url)
            try:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT dt.upload_id
                          FROM document_text dt
                          JOIN document_uploads du ON du.id = dt.upload_id
                         WHERE dt.tenant_id        = %s::uuid
                           AND dt.markdown_sha256  = %s
                           AND dt.upload_id       <> %s::uuid
                           AND du.extraction_status <> 'duplicate'
                         LIMIT 1
                        """,
                        (tenant_id, markdown_sha256, exclude_upload_id),
                    )
                    row = cur.fetchone()
                    return str(row[0]) if row else None
            finally:
                conn.close()
        except Exception as e:
            logger.warning(f"markdown dup-check failed: {e}")
            return None

    def _mark_duplicate(
        self,
        upload_id:        str,
        dup_of_upload_id: str,
        file_path:        str,
    ) -> None:
        """Mark upload as a duplicate of an existing row and delete the
        redundant file from disk. Best-effort — failures are logged."""
        try:
            import psycopg2
            conn = psycopg2.connect(self.db_url)
            try:
                with conn.cursor() as cur:
                    # series_id / version_no are nulled out so the row
                    # doesn't claim a slot in the visible version sequence
                    # (schema_v20). The dup-tombstone remains queryable by
                    # /status via upload_id, just not via /versions.
                    cur.execute(
                        """
                        UPDATE document_uploads
                           SET extraction_status = 'duplicate',
                               dup_of_upload_id  = %s::uuid,
                               processed_at      = now(),
                               error_message     = NULL,
                               findings_count    = 0,
                               series_id         = NULL,
                               version_no        = NULL
                         WHERE id = %s::uuid
                        """,
                        (dup_of_upload_id, upload_id),
                    )
                conn.commit()
            finally:
                conn.close()
        except Exception as e:
            logger.warning(f"mark duplicate failed for {upload_id}: {e}")

        try:
            from pathlib import Path as _P
            _p = _P(file_path)
            if _p.exists():
                _p.unlink()
        except Exception as e:
            logger.debug(f"duplicate file unlink failed for {file_path}: {e}")

    def _print_dry_run(self, findings, doc) -> None:
        print(f"\n{'='*60}")
        print(f"DRY RUN — {doc.original_name}")
        print(f"  doc_type:    {doc.doc_type}")
        print(f"  standards:   {doc.standard_ids}")
        print(f"  path:        {doc.extraction_path.value}")
        print(f"  tokens:      ~{doc.token_estimate:,}")
        print(f"  findings:    {len(findings)}")
        print()
        if findings:
            print("  Findings:")
            for f in sorted(findings, key=lambda x: x.control_ref):
                conf = f"[{f.confidence}]"
                evid = f.evidence_text[:60] + "..." if len(f.evidence_text) > 60 else f.evidence_text
                print(f"    {f.control_ref:12s} {f.finding:8s} {conf:8s} {evid}")
        print(f"{'='*60}\n")

    def _print_result(self, r: PipelineResult) -> None:
        status_icon = {"extracted": "✓", "failed": "✗", "manual_review": "△", "dry_run": "○"}.get(r.status, "?")
        print(
            f"  {status_icon} {r.document_name[:45]:45s} "
            f"{r.status:15s} "
            f"{r.findings_count:3d} findings  "
            f"{r.duration_ms:5d}ms"
        )
        if r.error:
            print(f"    ✗ {r.error[:80]}")

    def _print_summary(self, results: list[PipelineResult]) -> None:
        total    = len(results)
        ok       = sum(1 for r in results if r.status == "extracted")
        failed   = sum(1 for r in results if r.status == "failed")
        manual   = sum(1 for r in results if r.status == "manual_review")
        findings = sum(r.findings_count for r in results)
        print(f"\n{'='*60}")
        print(f"PIPELINE SUMMARY")
        print(f"  Documents: {total}  ✓ {ok}  ✗ {failed}  △ manual_review {manual}")
        print(f"  Findings:  {findings}")
        print(f"{'='*60}\n")


# =============================================================================
# CLI
# =============================================================================

def _resolve_original_filename(
    db_url: str, tenant_id: str, file_path: str,
) -> Optional[str]:
    """Best-effort lookup: when --file points at a stored upload's
    storage_path, return the user-facing filename from document_uploads.
    Returns None on any DB failure (caller falls back to on-disk basename).

    Closes the CLI <-> doc_mappings UX gap discovered 2026-06-08: the
    storage_path basename is a UUID (the API renames on upload), and the
    doc_mappings filename fingerprints all assume the user-facing name.
    Without this resolution, reprocessing an existing upload by file path
    skips every doc_mapping and silently falls back to the legacy broad
    DOC_TYPE_CLAUSE_MAP path."""
    if not db_url:
        return None
    try:
        import psycopg
        from pathlib import Path
        abs_path = str(Path(file_path).resolve())
        with psycopg.connect(db_url) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT set_config('app.tenant_id', %s, TRUE)",
                    (tenant_id,),
                )
                cur.execute(
                    """
                    SELECT filename FROM document_uploads
                     WHERE tenant_id    = %s
                       AND storage_path = %s
                     ORDER BY created_at DESC
                     LIMIT 1
                    """,
                    (tenant_id, abs_path),
                )
                row = cur.fetchone()
                if row and row[0]:
                    return row[0]
    except Exception:
        pass
    return None


def main():
    parser = argparse.ArgumentParser(
        description="ArionComply Document Injection Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 rag/intake/doc_pipeline.py --file policy.pdf --tenant-id <UUID>
  python3 rag/intake/doc_pipeline.py --dir ./docs/ --tenant-id <UUID> --dry-run
  python3 rag/intake/doc_pipeline.py --file report.xlsx --tenant-id <UUID> --trace
        """,
    )
    parser.add_argument("--file",      help="Path to a single document")
    parser.add_argument("--dir",       help="Directory of documents to process")
    parser.add_argument("--upload-id", help="Existing document_uploads.id to reprocess")
    parser.add_argument("--tenant-id", required=True, help="Tenant UUID")
    parser.add_argument("--dry-run",   action="store_true", help="Parse and extract without writing to DB")
    parser.add_argument("--verbose",   action="store_true", help="Debug logging")
    parser.add_argument("--trace",     action="store_true", help="Write trace rows to intake_trace_log")
    parser.add_argument("--output",    help="Write results to JSON file")
    parser.add_argument(
        "--original-name",
        help="User-facing filename override. The reader otherwise uses the on-disk basename, "
             "which is a UUID for API uploads — that breaks doc_mappings filename fingerprints. "
             "If omitted, the CLI auto-resolves from document_uploads.storage_path when --file "
             "points at a stored upload.",
    )
    args = parser.parse_args()

    if not args.file and not args.dir:
        parser.error("Provide --file or --dir")

    db_url  = os.getenv("DATABASE_URL")
    api_key = os.getenv("ANTHROPIC_API_KEY")

    if not db_url:
        parser.error("DATABASE_URL not set in environment")
    if not api_key:
        parser.error("ANTHROPIC_API_KEY not set in environment")

    pipeline = DocumentPipeline(
        db_url  = db_url,
        api_key = api_key,
        dry_run = args.dry_run,
        verbose = args.verbose,
        trace   = args.trace,
    )

    if args.file:
        # Resolve original_filename for doc_mappings filename matching.
        # Order: explicit --original-name > auto-resolve from document_uploads
        # (when --file points at a stored upload's storage_path) > the
        # on-disk basename (fallback; works for fresh local files but
        # produces UUID-named docs that don't match canonical fingerprints).
        original_name = args.original_name
        if not original_name:
            original_name = _resolve_original_filename(
                db_url, args.tenant_id, args.file,
            )
        result  = pipeline.run(
            args.file, args.tenant_id, args.upload_id,
            original_filename = original_name,
        )
        results = [result]
        pipeline._print_result(result)
    else:
        results = pipeline.run_directory(args.dir, args.tenant_id)

    if args.output:
        with open(args.output, "w") as f:
            json.dump(
                [
                    {
                        "upload_id":   r.upload_id,
                        "document":    r.document_name,
                        "status":      r.status,
                        "doc_type":    r.doc_type,
                        "standards":   r.standard_ids,
                        "path":        r.extraction_path,
                        "findings":    r.findings_count,
                        "duration_ms": r.duration_ms,
                        "error":       r.error,
                    }
                    for r in results
                ],
                f, indent=2,
            )
        print(f"Results written to {args.output}")

    failed = [r for r in results if r.status == "failed"]
    sys.exit(1 if failed else 0)


if __name__ == "__main__":
    main()
