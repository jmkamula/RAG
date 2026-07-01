"""
rag/templates/xlsx_renderer.py — render a tabular template as .xlsx.

For the 6 tabular/hybrid v2 anchors (A.5.9 Asset Inventory, 10.1
Improvement Action Register, Art.32 Risk Measures, 5.3 RACI,
6.1.3 SoA, Art.30 RoPA — and any future tabular leaves), Excel is
the native format: tenants can sort, filter, add rows in their
familiar tool. This module produces an .xlsx attachment from the
same data the markdown renderer uses.

Workbook structure:
  - "Register" sheet — the per-row table
      Row 1: column headers (frozen)
      Row 2+: pre-filled tenant data from tabular_evidence_rows
              (blank when no prior data)
  - "Guidance" sheet — per-column auditor-grade guidance pulled from
              the catalog (ChecklistItem.text for each MUST)
  - "_arion_meta" sheet (HIDDEN) — leaf_id + column → item_id mapping;
              foundation for future Phase B round-trip upload extractor.

The hidden _arion_meta sheet is a foundation move: an xlsx upload
extractor (not yet built) will read these named cells to recover
the binding from a tenant-saved workbook even if columns were
reordered / renamed. Phase A only writes them.
"""
from __future__ import annotations

import io
import re
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


_TABLE_COLUMNS_RE = re.compile(
    r"<!--\s*TABLE-COLUMNS\s+leaf:(req:[A-Za-z0-9.]+:[a-z0-9_]+)\s*-->"
    r"(.*?)"
    r"<!--\s*/TABLE-COLUMNS\s*-->",
    re.DOTALL | re.IGNORECASE,
)
_TABLE_COLUMN_RE = re.compile(
    r"<!--\s*column:\s*(item:[A-Za-z0-9.]+:[a-z0-9_]+)\s*-->",
)
_ITEM_EDIT_ZONE_RE = re.compile(
    r"<!--\s*EDIT-ZONE-START\s+(item:[A-Za-z0-9.]+:[a-z0-9_]+)\s*-->",
    re.IGNORECASE,
)


def _extract_table_columns(template_body: str, leaf_id: str) -> Optional[list[str]]:
    """Return ordered list of item_ids for the TABLE-COLUMNS metadata
    block matching this leaf_id. None if not tabular."""
    for m in _TABLE_COLUMNS_RE.finditer(template_body):
        if m.group(1) == leaf_id:
            cols = _TABLE_COLUMN_RE.findall(m.group(2))
            if cols:
                return cols
    return None


def _extract_doc_level_must_ids(
    template_body: str,
    column_ids:    list[str],
) -> list[str]:
    """For a hybrid template, return the item_ids that have per-MUST
    EDIT-ZONE markers but are NOT table columns. These are the
    document-level narrative MUSTs (e.g. owner, version,
    communicated, a52_consistency).

    Pure-tabular templates return []. The result preserves first-seen
    order in the body so the xlsx Document Fields sheet reads
    top-to-bottom matching the markdown structure.
    """
    column_set = set(column_ids)
    seen: set[str] = set()
    out:  list[str] = []
    for m in _ITEM_EDIT_ZONE_RE.finditer(template_body):
        mid = m.group(1)
        if mid in column_set or mid in seen:
            continue
        seen.add(mid)
        out.append(mid)
    return out


def _fetch_must_texts(pg_conn, item_ids: list[str]) -> dict[str, str]:
    """Pull MUST text per item_id from Neo4j ChecklistItem nodes.

    Uses the Neo4j driver since the catalog is loaded there for the
    engine. Falls back to empty dict on any error.
    """
    if not item_ids:
        return {}
    try:
        from rag.posture.advisory import _get_neo_driver
        driver = _get_neo_driver()
        if not driver:
            return _fallback_must_texts(item_ids)
        out: dict[str, str] = {}
        with driver.session() as session:
            result = session.run(
                "MATCH (ci:ChecklistItem) WHERE ci.id IN $ids "
                "RETURN ci.id AS id, ci.text AS text",
                ids=item_ids,
            )
            for row in result:
                out[row["id"]] = row["text"] or ""
        return out
    except Exception:
        return _fallback_must_texts(item_ids)


def _fallback_must_texts(item_ids: list[str]) -> dict[str, str]:
    """Fall back to the Python catalog when Neo4j is unavailable."""
    try:
        from enrichment.documents import document_requirements as drm
        out: dict[str, str] = {}
        wanted = set(item_ids)
        for attr in dir(drm):
            obj = getattr(drm, attr)
            if not isinstance(obj, drm.EvidenceRequirement):
                continue
            for ci in list(obj.must_contain) + list(obj.should_contain):
                if ci.id in wanted:
                    out[ci.id] = ci.text
        return out
    except Exception:
        return {}


def _fetch_prior_rows(pg_conn, tenant_id: str, leaf_id: str) -> list[dict]:
    """Pull active per-row content from tabular_evidence_rows for this
    tenant + leaf, in row_index order."""
    rows: list[dict] = []
    with pg_conn.cursor() as cur:
        cur.execute(
            "SELECT set_config('app.tenant_id', %s, TRUE)", (tenant_id,),
        )
        cur.execute(
            """
            SELECT row_index, column_values
              FROM tabular_evidence_rows
             WHERE tenant_id = %s::uuid
               AND leaf_id   = %s
               AND is_active = TRUE
             ORDER BY document_id, row_index
            """,
            (tenant_id, leaf_id),
        )
        for ri, cv in cur.fetchall():
            rows.append({"row_index": ri, "column_values": cv or {}})
    return rows


def _humanize_column(item_id: str) -> str:
    """Turn item:A.5.9:owner_per_asset → 'Owner Per Asset'.

    Used as a fallback column header when MUST text is unavailable
    or excessively long for an Excel header row.
    """
    parts = item_id.split(":")
    slug = parts[-1] if parts else item_id
    return slug.replace("_", " ").title()


def render_template_xlsx(
    pg_conn,
    tenant_id: str,
    leaf_id:   str,
    template_body: str,
) -> Optional[bytes]:
    """Generate an .xlsx workbook for a tabular template leaf.

    Returns the workbook bytes, or None if the template has no
    TABLE-COLUMNS metadata for this leaf (i.e. it's not a tabular
    template — caller should fall back to markdown).
    """
    column_ids = _extract_table_columns(template_body, leaf_id)
    if not column_ids:
        return None

    must_texts = _fetch_must_texts(pg_conn, column_ids)
    prior_rows = _fetch_prior_rows(pg_conn, tenant_id, leaf_id)

    # ── Workbook ─────────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1: Register
    ws = wb.active
    ws.title = "Register"

    # Header row: ALWAYS use the humanized slug — keeps headers compact
    # and consistent across columns. Full auditor-grade MUST text lives
    # on the Guidance sheet keyed by the same header label.
    headers = [_humanize_column(iid) for iid in column_ids]
    ws.append(headers)

    # Header style
    header_fill = PatternFill("solid", fgColor="3A382E")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(border_style="thin", color="C8C5B8")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col_ix, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_ix)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Pre-filled data rows
    body_border = Border(
        top=Side(border_style="thin", color="E8E5DA"),
        bottom=Side(border_style="thin", color="E8E5DA"),
        left=Side(border_style="thin", color="E8E5DA"),
        right=Side(border_style="thin", color="E8E5DA"),
    )
    for r in prior_rows:
        cv = r["column_values"]
        values = [cv.get(iid, "") for iid in column_ids]
        ws.append(values)
        for col_ix in range(1, len(values) + 1):
            cell = ws.cell(row=ws.max_row, column=col_ix)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = body_border

    # Append blank rows so the tenant has visible space to add new entries
    blank_to_add = max(0, 10 - len(prior_rows))
    for _ in range(blank_to_add):
        ws.append([""] * len(column_ids))
        for col_ix in range(1, len(column_ids) + 1):
            cell = ws.cell(row=ws.max_row, column=col_ix)
            cell.border = body_border

    # Column widths — proportional, capped
    for col_ix, header in enumerate(headers, start=1):
        # Heuristic: header length × 1.2, min 14, max 50
        width = max(14, min(50, int(len(header) * 1.2 + 6)))
        ws.column_dimensions[get_column_letter(col_ix)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Sheet 2: Guidance
    g = wb.create_sheet("Guidance")
    g.append(["Register column", "What auditors expect to see"])
    for col_ix in range(1, 3):
        cell = g.cell(row=1, column=col_ix)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = border
        cell.alignment = Alignment(vertical="center")
    for iid, header in zip(column_ids, headers):
        text = must_texts.get(iid, "") or "(no catalog text)"
        g.append([header, text])
        # Style data rows
        for col_ix in (1, 2):
            cell = g.cell(row=g.max_row, column=col_ix)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = body_border
    g.column_dimensions["A"].width = 22
    g.column_dimensions["B"].width = 90

    # Optional Sheet 3: Document Fields (hybrid templates only).
    # When the template carries doc-level narrative MUSTs (per-MUST
    # EDIT-ZONE markers outside the table — e.g. owner, version,
    # communicated, a52_consistency for 5.3 RACI), surface them as
    # key/value rows so the tenant can fill prose alongside the table
    # without switching to a .md or .docx file.
    doc_level_ids = _extract_doc_level_must_ids(template_body, column_ids)
    if doc_level_ids:
        # Add doc-level MUST texts to the catalog lookup if missing
        missing_texts = [mid for mid in doc_level_ids if mid not in must_texts]
        if missing_texts:
            must_texts.update(_fetch_must_texts(pg_conn, missing_texts))

        d = wb.create_sheet("Document Fields")
        d.append(["Field", "What auditors expect to see", "Your content"])
        # Style header
        for col_ix in range(1, 4):
            cell = d.cell(row=1, column=col_ix)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        for mid in doc_level_ids:
            label = _humanize_column(mid)
            stext = must_texts.get(mid, "(no catalog text)")
            d.append([label, stext, ""])
            for col_ix in (1, 2, 3):
                cell = d.cell(row=d.max_row, column=col_ix)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border    = body_border
            # Make the "Your content" cell tall enough for prose
            d.row_dimensions[d.max_row].height = 80
        d.column_dimensions["A"].width = 22
        d.column_dimensions["B"].width = 60
        d.column_dimensions["C"].width = 60
        d.freeze_panes = "A2"

    # HIDDEN sheet: metadata for future round-trip upload extractor
    m = wb.create_sheet("_arion_meta")
    m.append(["key", "value"])
    m.append(["leaf_id", leaf_id])
    m.append(["tenant_id", tenant_id])
    m.append(["column_count", len(column_ids)])
    for ix, iid in enumerate(column_ids):
        m.append([f"column_{ix:02d}", iid])
    if doc_level_ids:
        m.append(["doc_field_count", len(doc_level_ids)])
        for ix, mid in enumerate(doc_level_ids):
            m.append([f"doc_field_{ix:02d}", mid])
    m.sheet_state = "hidden"

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
