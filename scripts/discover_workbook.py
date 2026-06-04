"""CLI driver for Stage I workbook discovery.

Reads an .xlsx / .xlsm with openpyxl, runs the discovery engine against every
sheet, prints proposals, and (with --persist) writes them to
workbook_intake_proposal as a single discovery run.

Usage:
  python3 scripts/discover_workbook.py /path/to/workbook.xlsx
  python3 scripts/discover_workbook.py /path/to/workbook.xlsx --sheet "Asset Register"
  python3 scripts/discover_workbook.py /path/to/workbook.xlsx --json
  python3 scripts/discover_workbook.py /path/to/workbook.xlsx --persist \\
      --tenant-id 00000000-0000-0000-0000-000000000001
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict
from pathlib import Path
from uuid import UUID

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

import openpyxl  # noqa: E402

from rag.intake.workbook_discovery import discover_workbook  # noqa: E402


def _load_rows(workbook_path: Path, only_sheets: list[str] | None) -> dict[str, list[list]]:
    wb = openpyxl.load_workbook(workbook_path, keep_vba=True, data_only=True)
    out: dict[str, list[list]] = {}
    for name in wb.sheetnames:
        if only_sheets and name not in only_sheets:
            continue
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        out[name] = rows
    return out


def _print_proposal(p) -> None:
    print(f"\n━━ {p.sheet}  →  {p.mapping_id}  (confidence {p.confidence})")
    print(f"   path:       {Path(p.mapping_path).name}")
    print(f"   header_row: {p.header_row}    rows: {p.row_count}    headers: {len(p.headers)}")
    if p.warnings:
        for w in p.warnings:
            print(f"   ! {w}")
    for pp in p.passes:
        sat = len(pp.satisfied)
        par = len(pp.partial)
        mis = len(pp.missing)
        total = sat + par + mis
        verdict = "NC" if (sat + par) < total else "OFI" if par else "OK"
        print(f"   · pass {pp.pass_name!r:25s} target {pp.target_evidence_requirement}")
        print(f"     {sat}/{total} satisfied, {par} partial, {mis} missing  → engine-leaning {verdict}")
        if pp.freshness_column or pp.freshness_days:
            print(f"     freshness: col={pp.freshness_column!r}  days={pp.freshness_days}")
        if pp.warnings:
            for w in pp.warnings:
                print(f"     ! {w}")
        if pp.satisfied:
            print(f"     satisfied:")
            for mid in pp.satisfied:
                print(f"       ✓ {mid}  ← {pp.matched_columns.get(mid, '?')!r}")
        if pp.partial:
            print(f"     partial:")
            for mid in pp.partial:
                print(f"       ~ {mid}  ← {pp.matched_columns.get(mid, '?')!r}")
        if pp.missing:
            print(f"     missing:")
            for mid in pp.missing:
                print(f"       ✗ {mid}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("workbook", help="Path to .xlsx or .xlsm")
    ap.add_argument("--sheet", action="append", help="Limit to one or more sheet names (repeatable)")
    ap.add_argument("--floor", type=float, default=0.0, help="Confidence floor (0.0-1.0); proposals below are suppressed")
    ap.add_argument("--json", action="store_true", help="Emit JSON instead of human-readable text")
    ap.add_argument("--mappings-dir", default=None, help="Override mappings dir (default: db/workbook_mappings)")
    ap.add_argument("--persist", action="store_true", help="Write proposals to workbook_intake_proposal (requires --tenant-id)")
    ap.add_argument("--tenant-id", default=None, help="Tenant UUID for --persist")
    args = ap.parse_args()

    if args.persist and not args.tenant_id:
        print("--persist requires --tenant-id", file=sys.stderr)
        return 2

    wb_path = Path(args.workbook)
    if not wb_path.exists():
        print(f"workbook not found: {wb_path}", file=sys.stderr)
        return 2

    rows = _load_rows(wb_path, args.sheet)
    if not rows:
        print("no sheets read (filter may have excluded everything)", file=sys.stderr)
        return 2

    proposals = discover_workbook(
        rows,
        mappings_dir=Path(args.mappings_dir) if args.mappings_dir else None,
        confidence_floor=args.floor,
    )

    if args.json:
        print(json.dumps([asdict(p) for p in proposals], indent=2))
    else:
        print(f"workbook: {wb_path}")
        print(f"sheets read: {len(rows)}    proposals: {len(proposals)}")
        matched_sheets = {p.sheet for p in proposals}
        unmatched = [s for s in rows if s not in matched_sheets]
        if unmatched:
            print(f"unmatched sheets ({len(unmatched)}): {', '.join(unmatched)}")
        for p in proposals:
            _print_proposal(p)

    if args.persist:
        from rag.posture_loader import build_pg_conn
        from rag.intake.workbook_persistence import persist_proposals

        tenant_uuid = UUID(args.tenant_id)
        pg = build_pg_conn()
        try:
            run_id = persist_proposals(
                pg, tenant_uuid, str(wb_path), proposals
            )
        finally:
            pg.close()
        print(f"\npersisted {len(proposals)} proposal(s) as discovery_run_id={run_id}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
