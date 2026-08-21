#!/usr/bin/env python3
"""Ship 91'.f — hyperlink audit across unstructured.io capture.

Analyses every hyperlink + cross-sheet-formula-ref captured by
`readers.py::_partition_xlsx_via_unstructured` on a target workbook.
For each hyperlink, categorises the disposition:

  A. cite_emitted        — landed in a `cite_columns:` fingerprint match
                           on a sheet-fingerprint-matched YAML pass
  B. mailto_filtered     — dropped by row-level guard (Ship 89'.b)
  C. header_row          — hyperlink on the header row (not data)
  D. unmatched_column    — matched sheet + YAML has cite_columns but this
                           column doesn't fingerprint any of them
  E. no_cite_columns     — matched sheet + YAML has NO cite_columns block
                           (mapping is data-only or attestation)
  F. unmatched_sheet     — sheet doesn't fingerprint any YAML mapping

Also tallies cross-sheet formula refs (Ship 85'.a captures them
as `structured_sheets.cross_sheet_refs`) — those aren't hyperlinks
but they represent semantic external references worth counting.

Output: JSON + markdown summary. Feeds Ship 91'.g criterion doc.

Usage:
  POSTGRES_PASSWORD=... python scripts/ship91f_hyperlink_audit.py \\
    --workbook "ISO 27001 workbook Arion Networks.xlsm" \\
    --out /tmp/hyperlink_audit.md
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

sys.path.insert(0, "/data/arioncomply")

from rag.intake.readers import read_document
from rag.intake.workbook_discovery import (
    discover_workbook, load_mappings, tokenize, _find_column,
)


def _cell_column_letter(cell: str) -> str:
    out = []
    for ch in cell or "":
        if ch.isalpha(): out.append(ch)
        else: break
    return "".join(out).upper()


def _cell_row_number(cell: str) -> Optional[int]:
    digits = "".join(ch for ch in (cell or "") if ch.isdigit())
    if not digits: return None
    try: return int(digits)
    except ValueError: return None


def _column_letter_to_index(letters: str) -> int:
    idx = 0
    for ch in letters.upper():
        if not ch.isalpha(): break
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _classify_url(url: str) -> str:
    u = (url or "").lower()
    if u.startswith("mailto:"): return "mailto"
    if u.startswith(("http://", "https://")): return "http"
    if u.startswith(("../", "./")) or ":x:" in u or ":w:" in u or ":p:" in u:
        return "sharepoint_relative"
    if u.startswith("file:"): return "file"
    if u.startswith("#"): return "anchor"
    return "other"


def audit_workbook(workbook_path: str) -> dict:
    """Run the full audit — return a dict of stats + per-hyperlink rows."""
    doc = read_document(workbook_path)
    structured = doc.extraction_metrics.get("structured_sheets") or []
    if not structured:
        return {"error": "no structured_sheets captured — is this xlsx/xlsm?"}

    # Also load the raw row data for header detection + column-index lookup
    import openpyxl
    wb = openpyxl.load_workbook(workbook_path, keep_vba=True, data_only=True, read_only=True)
    rows_per_sheet = {sn: [list(r) for r in wb[sn].iter_rows(values_only=True)] for sn in wb.sheetnames}
    wb.close()

    proposals = discover_workbook(rows_per_sheet)
    proposals_by_sheet: dict[str, list] = defaultdict(list)
    for p in proposals:
        proposals_by_sheet[p.sheet].append(p)

    all_mappings = load_mappings()
    mapping_by_id = {m.get("mapping_id"): m for m in all_mappings}

    audit_rows: list[dict] = []
    tally: Counter = Counter()
    per_sheet: dict[str, Counter] = defaultdict(Counter)
    unmatched_columns_seen: Counter = Counter()  # (sheet, col_header) → count
    unmatched_sheets_seen: Counter = Counter()   # sheet → count
    cross_sheet_ref_count = 0

    for s in structured:
        sheet_name = s.get("sheet_name") or "?"
        hyperlinks = s.get("hyperlinks") or []
        xrefs = s.get("cross_sheet_refs") or []
        cross_sheet_ref_count += len(xrefs)
        if not hyperlinks:
            continue

        proposals_here = proposals_by_sheet.get(sheet_name) or []
        if not proposals_here:
            for h in hyperlinks:
                kind = _classify_url(h.get("url") or "")
                if kind == "mailto":
                    disposition = "B_mailto_filtered"
                else:
                    disposition = "F_unmatched_sheet"
                    unmatched_sheets_seen[sheet_name] += 1
                tally[disposition] += 1
                per_sheet[sheet_name][disposition] += 1
                audit_rows.append({
                    "sheet": sheet_name, "cell": h.get("cell"),
                    "url": (h.get("url") or "")[:120],
                    "url_kind": kind, "disposition": disposition,
                    "column_header": None, "matched_mapping": None,
                })
            continue

        # For each hyperlink: determine header of its column + whether it's
        # covered by any proposal's cite_columns block.
        for prop in proposals_here:
            mapping = mapping_by_id.get(prop.mapping_id) or {}
            passes = mapping.get("passes") or []
            headers = prop.headers or []
            header_row = prop.header_row
            has_cite_columns = any(p.get("cite_columns") for p in passes)
            # Precompute header text → column letter
            def _idx_to_letter(i: int) -> str:
                res = ""
                n = i + 1
                while n > 0:
                    n, r = divmod(n - 1, 26)
                    res = chr(ord("A") + r) + res
                return res
            # Header text → col index (first occurrence)
            header_col_idx: dict[str, int] = {}
            for i, h in enumerate(headers):
                if h and h not in header_col_idx:
                    header_col_idx[h] = i

            # Precompute: header letter → header text
            letter_to_header: dict[str, str] = {}
            for i, h in enumerate(headers):
                if h:
                    letter_to_header[_idx_to_letter(i)] = h

            # For each hyperlink, walk the passes' cite_columns to see if any match
            header_tokens = [tokenize(h) for h in headers] if headers else []
            for h in hyperlinks:
                url = h.get("url") or ""
                cell = h.get("cell") or ""
                kind = _classify_url(url)
                col_letter = _cell_column_letter(cell)
                row_num = _cell_row_number(cell)
                col_header = letter_to_header.get(col_letter)

                # (B) mailto guard applies regardless
                if kind == "mailto":
                    tally["B_mailto_filtered"] += 1
                    per_sheet[sheet_name]["B_mailto_filtered"] += 1
                    audit_rows.append({
                        "sheet": sheet_name, "cell": cell, "url": url[:120],
                        "url_kind": kind, "disposition": "B_mailto_filtered",
                        "column_header": col_header,
                        "matched_mapping": prop.mapping_id,
                    })
                    continue

                # (C) header-row hyperlink
                if header_row is not None and row_num is not None and row_num <= header_row + 1:
                    tally["C_header_row"] += 1
                    per_sheet[sheet_name]["C_header_row"] += 1
                    audit_rows.append({
                        "sheet": sheet_name, "cell": cell, "url": url[:120],
                        "url_kind": kind, "disposition": "C_header_row",
                        "column_header": col_header,
                        "matched_mapping": prop.mapping_id,
                    })
                    continue

                # (E) mapping has no cite_columns block at all
                if not has_cite_columns:
                    tally["E_no_cite_columns"] += 1
                    per_sheet[sheet_name]["E_no_cite_columns"] += 1
                    audit_rows.append({
                        "sheet": sheet_name, "cell": cell, "url": url[:120],
                        "url_kind": kind, "disposition": "E_no_cite_columns",
                        "column_header": col_header,
                        "matched_mapping": prop.mapping_id,
                    })
                    continue

                # (A/D) check if the column fingerprint-matches any pass's
                # cite_columns fingerprint
                matched_binding = None
                for pass_yaml in passes:
                    for cite_col in pass_yaml.get("cite_columns") or []:
                        fp = cite_col.get("fingerprint") or []
                        alt = cite_col.get("alternative_fingerprints") or None
                        hit = _find_column(fp, alt, header_tokens, headers)
                        if hit is None:
                            continue
                        _, matched_header = hit
                        if matched_header == col_header:
                            matched_binding = {
                                "pass_name": pass_yaml.get("pass_name"),
                                "binds_to":  cite_col.get("binds_to"),
                                "cite_kind": cite_col.get("cite_kind", "internal_document"),
                            }
                            break
                    if matched_binding:
                        break

                if matched_binding:
                    tally["A_cite_emitted"] += 1
                    per_sheet[sheet_name]["A_cite_emitted"] += 1
                    audit_rows.append({
                        "sheet": sheet_name, "cell": cell, "url": url[:120],
                        "url_kind": kind, "disposition": "A_cite_emitted",
                        "column_header": col_header,
                        "matched_mapping": prop.mapping_id,
                        "bound_must": matched_binding["binds_to"],
                    })
                else:
                    tally["D_unmatched_column"] += 1
                    per_sheet[sheet_name]["D_unmatched_column"] += 1
                    unmatched_columns_seen[(sheet_name, col_header or "?")] += 1
                    audit_rows.append({
                        "sheet": sheet_name, "cell": cell, "url": url[:120],
                        "url_kind": kind, "disposition": "D_unmatched_column",
                        "column_header": col_header,
                        "matched_mapping": prop.mapping_id,
                    })
            break  # Only count each hyperlink once even if multiple proposals per sheet

    total_hyperlinks = sum(tally[k] for k in tally if not k.startswith("_"))
    return {
        "workbook_path":            workbook_path,
        "total_hyperlinks":         total_hyperlinks,
        "total_cross_sheet_refs":   cross_sheet_ref_count,
        "disposition_tally":        dict(tally),
        "per_sheet_tally":          {sn: dict(c) for sn, c in per_sheet.items()},
        "unmatched_columns":        {f"{s} :: {c}": n for (s, c), n in unmatched_columns_seen.most_common()},
        "unmatched_sheets":         dict(unmatched_sheets_seen),
        "audit_rows":               audit_rows,
    }


def render_markdown(result: dict) -> str:
    if "error" in result:
        return f"# Ship 91'.f Hyperlink Audit\n\nError: {result['error']}\n"

    tally = result["disposition_tally"]
    total = result["total_hyperlinks"]
    lines = [
        "# Ship 91'.f — Hyperlink audit",
        "",
        f"**Workbook**: `{result['workbook_path']}`",
        "",
        f"**Total hyperlinks captured**: {total}",
        f"**Total cross-sheet formula refs**: {result['total_cross_sheet_refs']}",
        "",
        "## Disposition tally",
        "",
        "| Bucket | Count | % | Meaning |",
        "|---|---|---|---|",
    ]
    labels = {
        "A_cite_emitted":       "matched cite_columns → cite emitted",
        "B_mailto_filtered":    "mailto: dropped by row-level guard",
        "C_header_row":         "header-row hyperlink (metadata, not data)",
        "D_unmatched_column":   "mapping has cite_columns but this column doesn't match any fingerprint",
        "E_no_cite_columns":    "mapping has NO cite_columns block (data-only or attestation)",
        "F_unmatched_sheet":    "sheet doesn't fingerprint-match any mapping",
    }
    for key, label in labels.items():
        n = tally.get(key, 0)
        pct = f"{100*n/total:.1f}%" if total else "-"
        lines.append(f"| **{key}** | {n} | {pct} | {label} |")

    lines += ["", "## Per-sheet breakdown", ""]
    lines.append("| Sheet | Total | A | B | C | D | E | F |")
    lines.append("|---|---|---|---|---|---|---|---|")
    for sn, c in result["per_sheet_tally"].items():
        total_sheet = sum(c.values())
        lines.append(f"| {sn} | {total_sheet} | "
                     f"{c.get('A_cite_emitted', 0)} | "
                     f"{c.get('B_mailto_filtered', 0)} | "
                     f"{c.get('C_header_row', 0)} | "
                     f"{c.get('D_unmatched_column', 0)} | "
                     f"{c.get('E_no_cite_columns', 0)} | "
                     f"{c.get('F_unmatched_sheet', 0)} |")

    unmatched_cols = result.get("unmatched_columns") or {}
    if unmatched_cols:
        lines += ["", "## Unmatched columns (loose ends — catalog gaps to consider)", ""]
        lines.append("| Sheet :: Column | Count |")
        lines.append("|---|---|")
        for k, n in unmatched_cols.items():
            lines.append(f"| {k} | {n} |")

    unmatched_sheets = result.get("unmatched_sheets") or {}
    if unmatched_sheets:
        lines += ["", "## Unmatched sheets (no YAML fingerprint hit)", ""]
        lines.append("| Sheet | Hyperlinks |")
        lines.append("|---|---|")
        for sn, n in unmatched_sheets.items():
            lines.append(f"| {sn} | {n} |")

    lines += ["", "## Sample audit rows (first 20)", ""]
    lines.append("| Sheet | Cell | Column | Kind | Disposition | Bound MUST |")
    lines.append("|---|---|---|---|---|---|")
    for r in result["audit_rows"][:20]:
        lines.append(f"| {r['sheet']} | {r['cell']} | "
                     f"{r.get('column_header') or '?'} | "
                     f"{r['url_kind']} | {r['disposition']} | "
                     f"{r.get('bound_must') or '-'} |")
    return "\n".join(lines) + "\n"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True,
                    help="Basename OR full path. Basename resolves via document_uploads.")
    ap.add_argument("--out", help="Write markdown report to this path")
    ap.add_argument("--json", help="Also emit JSON at this path")
    args = ap.parse_args()

    # Resolve path
    path = args.workbook
    if not Path(path).exists():
        import psycopg2
        conn = psycopg2.connect(
            host="127.0.0.1", dbname="arioncomply_compliance",
            user="arioncomply",
            password=os.getenv("POSTGRES_PASSWORD", "arionlocal2026"),
        )
        with conn.cursor() as cur:
            cur.execute(
                "SELECT storage_path FROM document_uploads "
                "WHERE filename = %s LIMIT 1", (args.workbook,),
            )
            row = cur.fetchone()
        conn.close()
        if not row or not row[0]:
            print(f"Workbook not found: {args.workbook}", file=sys.stderr)
            return 2
        path = row[0]

    result = audit_workbook(path)
    md = render_markdown(result)
    if args.out:
        Path(args.out).write_text(md)
        print(f"Wrote report → {args.out}")
    else:
        print(md)
    if args.json:
        Path(args.json).write_text(json.dumps(result, indent=2, default=str))
        print(f"Wrote JSON → {args.json}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
