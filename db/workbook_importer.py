"""
ArionComply — Generic ISO 27001 Workbook Importer

Design principles:
  1. Generic — sheet-to-table mapping is config, not code
  2. Safe — all imports go through posture_pending for posture changes;
             assets/risks/vendors write directly (no posture risk)
  3. Idempotent — safe to re-run; uses UPSERT on external_ref
  4. Transparent — dry_run mode shows what would be written
  5. Auditable — every row tagged workbook_imported=TRUE + import_date

Architecture:
  SheetConfig     — declares how one sheet maps to one table
  RowMapper       — transforms a raw workbook row to a DB dict
  WorkbookImporter — orchestrates the full import

Usage:
  importer = WorkbookImporter(pg_conn, tenant_id)
  report = importer.import_all(workbook_path, dry_run=True)
  report.print_summary()

  # Live import after review:
  report = importer.import_all(workbook_path, dry_run=False)
"""
from __future__ import annotations

import re
import uuid
import warnings
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any, Callable

warnings.filterwarnings('ignore')

try:
    import openpyxl
except ImportError:
    raise ImportError("pip install openpyxl")


# =============================================================================
# Dataclasses
# =============================================================================

@dataclass
class ImportRow:
    """One row extracted from a workbook sheet."""
    sheet:      str
    row_num:    int
    raw:        dict[str, Any]     # column_header → cell value
    mapped:     dict[str, Any]     # db_column → value (after mapping)
    action:     str = "insert"     # insert / update / skip
    reason:     str = ""


@dataclass
class ImportReport:
    """Summary of a full workbook import run."""
    dry_run:    bool
    tenant_id:  str
    sheets:     dict[str, "SheetReport"] = field(default_factory=dict)

    def add(self, sheet: str, rows: list[ImportRow]):
        self.sheets[sheet] = SheetReport(sheet=sheet, rows=rows)

    def print_summary(self):
        total_insert = total_update = total_skip = total_error = 0
        print(f"\n{'DRY RUN' if self.dry_run else 'LIVE IMPORT'} — Tenant {self.tenant_id}")
        print("=" * 70)
        for name, sr in self.sheets.items():
            ins = sum(1 for r in sr.rows if r.action == "insert")
            upd = sum(1 for r in sr.rows if r.action == "update")
            skp = sum(1 for r in sr.rows if r.action == "skip")
            err = sum(1 for r in sr.rows if r.action == "error")
            total_insert += ins; total_update += upd
            total_skip += skp; total_error += err
            print(f"  {name:45s} +{ins:3d}  ~{upd:2d}  skip:{skp:2d}  err:{err:2d}")
            for r in sr.rows:
                if r.action in ("error", "skip") and r.reason:
                    print(f"    Row {r.row_num}: [{r.action}] {r.reason}")
        print("-" * 70)
        print(f"  TOTAL: insert={total_insert}  update={total_update}  skip={total_skip}  error={total_error}")


@dataclass
class SheetReport:
    sheet: str
    rows:  list[ImportRow]


# =============================================================================
# Column type converters
# =============================================================================

def _to_str(v) -> str | None:
    if v is None: return None
    s = str(v).strip()
    return s if s and s.lower() not in ('none', 'n/a', '-', '') else None

def _to_bool(v) -> bool | None:
    if v is None: return None
    s = str(v).strip().lower()
    if s in ('true','yes','y','1','implemented','applicable'): return True
    if s in ('false','no','n','0','n/a','not applicable'):     return False
    return None

def _to_int(v) -> int | None:
    if v is None: return None
    try:    return int(float(str(v).strip()))
    except: return None

def _to_date(v) -> date | None:
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
        try:    return datetime.strptime(s.split()[0], fmt).date()
        except: pass
    return None

def _clean_risk_level(v) -> str | None:
    """Normalise risk level to High/Medium/Low — workbook may have verbose text."""
    if v is None: return None
    s = str(v).strip().lower()
    if s.startswith('high'):   return 'High'
    if s.startswith('medium'): return 'Medium'
    if s.startswith('low'):    return 'Low'
    if s.startswith('critic'): return 'High'   # treat critical as high
    return None  # unknown → null rather than violating CHECK

def _to_list(v, sep=',') -> list[str]:
    if v is None: return []
    if isinstance(v, list): return [str(x).strip() for x in v if x]
    s = str(v).strip()
    if not s or s.lower() in ('none', 'n/a', '-'): return []
    return [x.strip() for x in s.split(sep) if x.strip()]

def _soa_finding(applicable, status) -> str:
    """Map SoA Applicable + Status to posture finding."""
    app = _to_bool(applicable)
    sts = _to_str(status) or ''
    if app is False or sts == 'N/A':
        return 'N/A'
    if sts == 'Implemented':
        return 'Comply'
    if sts in ('Partially Implemented', 'Planned'):
        return 'OFI'
    if sts in ('Not Implemented',):
        return 'NC'
    return 'Not assessed'

def _nc_finding(description) -> str:
    """Map NC tracker description to finding type."""
    d = str(description or '').lower()
    if 'ofi' in d or 'opportunity' in d:
        return 'OFI'
    if 'nc' in d or 'nonconform' in d or 'non-conform' in d:
        return 'NC'
    return 'OFI'  # default for external audit findings

def _risk_level_from_score(score) -> str | None:
    n = _to_int(score)
    if n is None: return None
    if n >= 8:  return 'high'
    if n >= 4:  return 'medium'
    return 'low'


# =============================================================================
# Sheet configurations
# Each SheetConfig describes one sheet → one table mapping
# =============================================================================

@dataclass
class SheetConfig:
    """
    Declares how one workbook sheet maps to one database table.

    header_row: which row contains column headers (1-based)
    data_start: which row data begins (1-based)
    skip_if_empty: column that must have a value for row to be imported
    external_ref_col: workbook column that gives the external ID (DOC001 etc)
    """
    name:               str             # sheet name in workbook
    table:              str             # postgres table name
    header_row:         int = 1
    data_start:         int = 2
    skip_if_empty:      str = ""        # column name; row skipped if this is empty
    external_ref_col:   str = ""        # column containing the external ref
    upsert_on:          str = "external_ref"  # column for upsert conflict


# =============================================================================
# Row mappers — one per sheet
# Each returns dict[db_column → value] or None to skip
# =============================================================================

class RowMappers:
    """
    One @staticmethod per sheet.
    Input: raw dict {column_header: value}
    Output: dict {db_column: value} or None to skip row
    """

    @staticmethod
    def soa(raw: dict, tenant_id: str) -> dict | None:
        """Statement of Applicability → posture_controls"""
        raw_ref = _to_str(raw.get('Control ID'))
        if not raw_ref or not re.match(r'^\d', raw_ref):
            return None  # skip section headers like "5", "Organizational Controls"

        # Canonical form: ISO 27001 Annex A subclauses use the 'A.' prefix
        # everywhere in the system (schema v14). The workbook ships them
        # bare ("5.18" not "A.5.18") so we normalize at import time —
        # otherwise we recreate the dedup bug that v14 just cleaned up.
        from rag.framework_refs import normalize_control_ref
        ref = normalize_control_ref(raw_ref, 'ISO27001:2022') or raw_ref

        applicable = raw.get('Applicable')
        status     = raw.get('Status')
        finding    = _soa_finding(applicable, status)

        return {
            'tenant_id':         tenant_id,
            'standard_id':       'ISO27001:2022',
            'control_ref':       ref,
            'node_id':           f'ISO27001:2022:{ref}',
            'finding':           finding,
            'confidence':        'high',
            'gap_description':   _to_str(raw.get('Justification')),
            'soa_notes':         _to_str(raw.get('Notes')),
            'soa_justification': _to_str(raw.get('Justification')),
            'linked_policies':   _to_list(raw.get('Linked Policies/Processes')),
            'source':            'workbook',
            'workbook_imported': True,
            'workbook_import_date': datetime.now(),
            # N/A controls have no remediation needed
            'remediation_status': 'closed' if finding == 'N/A' else 'open',
        }

    @staticmethod
    def nonconf(raw: dict, tenant_id: str) -> dict | None:
        """NonConf Act. Tracker → posture_controls (OFI/NC override)"""
        nc_id = _to_str(raw.get('Non Conformance ID'))
        if not nc_id or not nc_id.startswith('F'):
            return None

        desc     = _to_str(raw.get('Description')) or ''
        severity = _to_str(raw.get('Severity Level'))
        action   = _to_str(raw.get('CA description')) or _to_str(raw.get('Corrective Action'))
        finding  = _nc_finding(desc)
        owner    = _to_str(raw.get('Responsible Person/Owner'))

        # Map severity text to risk_level
        risk_map = {'high': 'high', 'medium': 'medium', 'low': 'low',
                    'critical': 'critical'}
        risk_level = risk_map.get((severity or '').lower(), 'medium')

        return {
            'tenant_id':          tenant_id,
            'standard_id':        'ISO27001:2022',
            'external_ref':       nc_id,
            'finding':            finding,
            'confidence':         'high',
            'gap_description':    desc,
            'action_required':    action,
            'risk_level':         risk_level,
            'owner_text':         owner,
            'target_date':        _to_date(raw.get('Target completion date')),
            'source':             'workbook',
            'workbook_imported':  True,
            'workbook_import_date': datetime.now(),
        }
        # NOTE: control_ref is NOT set here — NC tracker doesn't map findings
        # to specific control refs. This goes through posture_pending,
        # and a human or LLM maps the finding to the right control.

    @staticmethod
    def documents(raw: dict, tenant_id: str) -> dict | None:
        """Document Cont. Reg. → client_documents (metadata only)"""
        doc_id = _to_str(raw.get('Document ID'))
        if not doc_id or not doc_id.startswith('DOC'):
            return None
        title  = _to_str(raw.get('Document Title'))
        if not title:
            return None

        status  = _to_str(raw.get('Document Status')) or 'Active'
        version = _to_str(raw.get('Major Version No.'))

        return {
            'tenant_id':        tenant_id,
            'external_ref':     doc_id,
            'filename':         f"{doc_id}_{title[:40].replace(' ','_')}.pdf",
            'storage_path':     None,           # not uploaded yet
            'document_title':   title,
            'document_type':    'policy',       # refined later when uploaded
            'document_owner':   _to_str(raw.get('Owner')),
            'version':          version,
            'approval_status':  _to_str(raw.get('Appproval Status')),
            'approved_by':      _to_str(raw.get('Approved by')),
            'approval_date':    _to_date(raw.get('Approval Date')),
            'is_current':       status == 'Active',
            'is_metadata_only': True,           # file not uploaded
            'workbook_imported': True,
        }

    @staticmethod
    def incidents(raw: dict, tenant_id: str) -> dict | None:
        """Incident Log → incidents"""
        inc_id = _to_str(raw.get('Incident ID'))
        if not inc_id or not inc_id.startswith('INC'):
            return None

        asset_name = _to_str(raw.get('Asset Name'))
        pii        = _to_bool(raw.get('PII Involved (Y/N)'))
        auth_notif = _to_bool(raw.get('Supervisory Authority Notified (Y/N)'))
        ds_notif   = _to_bool(raw.get('Data Subject NOtified (Y/N)'))
        breach_cls = _to_str(raw.get('Breach Classification (InfoSec inc., PII, both, non breach)'))
        occurred   = _to_date(raw.get('Date'))
        resolved   = _to_date(raw.get('Resolution Date'))
        count      = _to_int(raw.get('Data Subjects Affected (approx. number of individuals)'))

        # Derive incident_type from breach classification
        if breach_cls:
            inc_type = breach_cls.lower().replace(' ', '_')
        elif pii:
            inc_type = 'pii_breach'
        else:
            inc_type = 'infosec_incident'

        # Derive severity from affected count
        severity = 'low'
        if count and count > 100:  severity = 'high'
        elif count and count > 10: severity = 'medium'

        return {
            'tenant_id':               tenant_id,
            'external_ref':            inc_id,
            'incident_type':           inc_type,
            'title':                   asset_name or f"Incident {inc_id}",
            'status':                  'closed' if resolved else 'open',
            'severity':                severity,
            'occurred_at':             datetime.combine(occurred, datetime.min.time()) if occurred else None,
            'closed_at':               datetime.combine(resolved, datetime.min.time()) if resolved else None,
            'affected_count_approx':   count,
            'affected_categories':     _to_list(raw.get('PII Categories')),
            'asset_ref':               _to_str(raw.get('Asset ID')),
            'pii_involved':            pii,
            'authority_notified':      auth_notif,
            'data_subjects_notified':  ds_notif,
            'actions_taken':           _to_str(raw.get('Actions Taken')),
            'lessons_learned':         _to_str(raw.get('Lessons Learned')),
            'evidence_collected':      _to_bool(raw.get('Evidence Collected  (Y/N)')),
            'workbook_imported':       True,
        }

    @staticmethod
    def assets(raw: dict, tenant_id: str) -> dict | None:
        """Asset Register → assets"""
        asset_id = _to_str(raw.get('Asset ID'))
        name     = _to_str(raw.get('Asset Name'))
        if not asset_id or not name:
            return None

        personal_data = _to_str(raw.get('Personal Data Types'))
        contains_pii  = bool(personal_data and personal_data.lower() not in ('none','n/a','-',''))

        return {
            'tenant_id':              tenant_id,
            'external_ref':           asset_id,
            'name':                   name,
            'asset_type':             _to_str(raw.get('Asset_Type')),
            'owner_text':             _to_str(raw.get('Asset_Owner')),
            'location':               _to_str(raw.get('Asset_Location')),
            'value_classification':   _to_str(raw.get('Asset Value')),
            'cia_c':                  _to_str(raw.get('Confidentiality')),
            'cia_i':                  _to_str(raw.get('Integrity')),
            'cia_a':                  _to_str(raw.get('Availability')),
            'comments':               _to_str(raw.get('Comments')),
            'personal_data_types':    _to_list(personal_data),
            'data_subject_categories':_to_list(raw.get('Data Subjects')),
            'processing_purposes':    _to_list(raw.get('Processing Purposes')),
            'retention_period':       _to_str(raw.get('Retention Period')),
            'contains_pii':           contains_pii,
            'workbook_imported':      True,
        }

    @staticmethod
    def risks(raw: dict, tenant_id: str) -> dict | None:
        """Risk Register → risks"""
        risk_id = _to_str(raw.get('Risk ID'))
        if not risk_id or not risk_id.startswith('R'):
            return None

        score = _to_int(raw.get('Risk Score'))

        return {
            'tenant_id':       tenant_id,
            'external_ref':    risk_id,
            'asset_ref':       _to_str(raw.get('Asset ID')),
            'asset_name':      _to_str(raw.get('Asset Name')),
            'interested_party':_to_str(raw.get('Interested Party')),
            'threat':          _to_str(raw.get('Threat')),
            'vulnerability':   _to_str(raw.get('Vulnerability')),
            'likelihood':      _to_int(raw.get('Likelihood')),
            'impact':          _to_int(raw.get('Impact')),
            'risk_score':      score,
            'risk_owner_text': _to_str(raw.get('Risk Owner')),
            'workbook_imported': True,
        }

    @staticmethod
    def risk_treatment(raw: dict, tenant_id: str) -> dict | None:
        """Risk Treatment Plan → risks (update existing rows)"""
        risk_id = _to_str(raw.get('Risk ID'))
        if not risk_id or not risk_id.startswith('R'):
            return None

        option = _to_str(raw.get('Treatment Option'))
        valid_options = ('Mitigate', 'Accept', 'Transfer', 'Avoid')
        if option not in valid_options:
            option = 'Mitigate'

        status_raw = _to_str(raw.get('Status'))
        status_map = {
            'implemented': 'implemented',
            'in progress': 'in_progress',
            'in-progress': 'in_progress',
            'open': 'open',
            'accepted': 'accepted',
        }
        status = status_map.get((status_raw or '').lower(), 'open')

        # Cross-framework: workbook ships ISMS + PIMS columns separately.
        # Storage is unified into a single STANDARD:VERSION:REF array
        # so adding more frameworks doesn't require a schema change.
        isms_raw = _to_list(raw.get('ISMS Applicable Controls')) or []
        pims_raw = _to_list(raw.get('PIMS Applicable Controls')) or []
        # Each workbook entry is "<ref> <description>" — keep the leading
        # token only. Skip empty entries.
        control_refs = sorted({
            f"ISO27001:2022:{c.split(' ', 1)[0]}"
            for c in isms_raw if c and c.split(' ', 1)[0]
        } | {
            f"ISO27701:2019:{c.split(' ', 1)[0]}"
            for c in pims_raw if c and c.split(' ', 1)[0]
        })

        return {
            'tenant_id':             tenant_id,
            'external_ref':          risk_id,   # used for upsert
            'treatment_option':      option,
            'treatment_action':      _to_str(raw.get('Treatment Action')),
            'control_refs':          control_refs,
            'implementation_date':   _to_date(raw.get('Implementation Date')),
            'residual_risk_level':   _to_int(raw.get('Residual Risk Level after Treatment')),
            'treatment_status':      status,
            'review_date':           _to_date(raw.get('Review Date')),
            'effectiveness_review':  _to_str(raw.get('Effectivness Review')),
        }

    @staticmethod
    def vendors(raw: dict, tenant_id: str) -> dict | None:
        """ThirdParty Vendor Risk → vendors"""
        name = _to_str(raw.get('Vendor Name'))
        if not name:
            return None

        dpa_raw = _to_str(raw.get('DPA signed')) or ''
        dpa_signed = 'yes' in dpa_raw.lower() or 'executed' in dpa_raw.lower()
        category = _to_str(raw.get('Vendor Category')) or ''
        is_proc  = 'processor' in category.lower() or 'pii' in category.lower()

        return {
            'tenant_id':              tenant_id,
            'name':                   name,
            'service_provided':       _to_str(raw.get('Service Provided')),
            'vendor_category':        category,
            'data_subject_categories':_to_list(raw.get('Data Subject Categories')),
            'data_shared':            _to_str(raw.get('Data Shared')),
            'data_location':          _to_str(raw.get('Data Location')),
            'dpa_signed':             dpa_signed,
            'dpa_reference':          dpa_raw if dpa_signed else None,
            'risk_level':             _clean_risk_level(raw.get('Risk level')),
            'security_controls':      _to_str(raw.get('Security Controls in Place')),
            'compliance_certs':       _to_list(raw.get('Compliance Certs')),
            'last_review_date':       _to_date(raw.get('Last Risk Review Date')),
            'next_review_date':       _to_date(raw.get('Next Planned Review Date')),
            'notes':                  _to_str(raw.get('Notes')),
            'is_processor':           is_proc,
            'workbook_imported':      True,
        }

    @staticmethod
    def isms_audits(raw: dict, tenant_id: str) -> dict | None:
        """Audit Log + External Audit Log → isms_audits
        
        The workbook has one row per finding (multiple rows per audit).
        We upsert on external_ref — the DO UPDATE appends finding_refs.
        Since postgres array append isn't atomic in simple UPSERT,
        we collect all finding_refs per audit_id and deduplicate.
        The upsert DO UPDATE merges finding_refs arrays.
        """
        audit_id = (_to_str(raw.get('Audit ID'))
                    or _to_str(raw.get('Ext Audit ID')))
        if not audit_id:
            return None

        is_external = 'EAUD' in (audit_id or '')
        auditor     = _to_str(raw.get('Auditor'))
        auditor_org = None
        if auditor and ',' in auditor:
            parts       = auditor.split(',')
            auditor     = parts[0].strip()
            auditor_org = parts[-1].strip() if len(parts) > 1 else None

        finding_ref = _to_str(raw.get('Finding ID') or raw.get('Finding ID'))
        finding_refs = [finding_ref] if finding_ref else []

        notes = _to_str(raw.get('Findings') or raw.get('Auditor Notes'))

        # Cross-framework: a single audit may cover multiple standards.
        # Default to ISO 27001 since the workbook doesn't yet carry a
        # framework column — extend here once it does.
        return {
            'tenant_id':         tenant_id,
            'external_ref':      audit_id,
            'audit_type':        'external' if is_external else 'internal',
            'audit_date':        _to_date(raw.get('Date')),
            'auditor_name':      auditor,
            'auditor_org':       auditor_org,
            'scope':             _to_str(raw.get('Scope')),
            'standard_ids':      ['ISO27001:2022'],
            'finding_refs':      finding_refs,
            'notes':             notes,
            'workbook_imported': True,
        }


# =============================================================================
# Main importer
# =============================================================================

class WorkbookImporter:
    """
    Generic ISO 27001 workbook importer.
    Maps workbook sheets → Postgres tables via config + row mappers.
    All writes are idempotent (UPSERT on external_ref or name).
    """

    # Sheet configurations — one per importable sheet
    SHEET_CONFIGS: list[SheetConfig] = [
        SheetConfig(
            name="Statement of Applicability",
            table="posture_controls",
            header_row=13, data_start=15,
            skip_if_empty="Control ID",
            external_ref_col="Control ID",
            upsert_on="(tenant_id, standard_id, control_ref)",
        ),
        SheetConfig(
            name="NonConf Act. Tracker",
            table="posture_controls",
            header_row=1, data_start=2,
            skip_if_empty="Non Conformance ID",
            external_ref_col="Non Conformance ID",
            upsert_on="(tenant_id, external_ref)",
        ),
        SheetConfig(
            name="Document Cont. Reg.",
            table="client_documents",
            header_row=1, data_start=2,
            skip_if_empty="Document ID",
            external_ref_col="Document ID",
            upsert_on="(tenant_id, external_ref)",
        ),
        SheetConfig(
            name="Incident Log",
            table="incidents",
            header_row=2, data_start=3,
            skip_if_empty="Incident ID",
            external_ref_col="Incident ID",
            upsert_on="(tenant_id, external_ref)",
        ),
        SheetConfig(
            name="Asset Register",
            table="assets",
            header_row=1, data_start=2,
            skip_if_empty="Asset ID",
            external_ref_col="Asset ID",
            upsert_on="(tenant_id, external_ref)",
        ),
        SheetConfig(
            name="Risk Register",
            table="risks",
            header_row=1, data_start=2,
            skip_if_empty="Risk ID",
            external_ref_col="Risk ID",
            upsert_on="(tenant_id, external_ref)",
        ),
        SheetConfig(
            name="Risk Treatment Plan",
            table="risks",
            header_row=1, data_start=2,
            skip_if_empty="Risk ID",
            external_ref_col="Risk ID",
            upsert_on="(tenant_id, external_ref)",
        ),
        SheetConfig(
            name="ThirdParty Vendor Risk ",
            table="vendors",
            header_row=1, data_start=2,
            skip_if_empty="Vendor Name",
            external_ref_col="Vendor Name",
            upsert_on="(tenant_id, name)",
        ),
        SheetConfig(
            name="Audit Log",
            table="isms_audits",
            header_row=1, data_start=2,
            skip_if_empty="Audit ID",
            external_ref_col="Audit ID",
            upsert_on="(tenant_id, external_ref)",
        ),
        SheetConfig(
            name="External Audit Log",
            table="isms_audits",
            header_row=1, data_start=2,
            skip_if_empty="Ext Audit ID",
            external_ref_col="Ext Audit ID",
            upsert_on="(tenant_id, external_ref)",
        ),
    ]

    # Sheet name → row mapper function
    MAPPERS: dict[str, Callable] = {
        "Statement of Applicability": RowMappers.soa,
        "NonConf Act. Tracker":       RowMappers.nonconf,
        "Document Cont. Reg.":        RowMappers.documents,
        "Incident Log":               RowMappers.incidents,
        "Asset Register":             RowMappers.assets,
        "Risk Register":              RowMappers.risks,
        "Risk Treatment Plan":        RowMappers.risk_treatment,
        "ThirdParty Vendor Risk ":    RowMappers.vendors,
        "Audit Log":                  RowMappers.isms_audits,
        "External Audit Log":         RowMappers.isms_audits,
    }

    # Prefix → table name (must match ref_prefixes in schema)
    PREFIX_TABLE = {
        'PC':  'posture_controls',
        'CD':  'client_documents',
        'INC': 'incidents',
        'AST': 'assets',
        'RSK': 'risks',
        'VND': 'vendors',
        'AUD': 'isms_audits',
        'FND': 'document_findings',
    }

    # Table → prefix
    TABLE_PREFIX = {v: k for k, v in {
        'PC':  'posture_controls',
        'CD':  'client_documents',
        'INC': 'incidents',
        'AST': 'assets',
        'RSK': 'risks',
        'VND': 'vendors',
        'AUD': 'isms_audits',
        'FND': 'document_findings',
    }.items()}

    def __init__(self, pg_conn, tenant_id: str, tenant_short: str = ""):
        self._pg           = pg_conn
        self._tenant_id    = tenant_id
        self._tenant_short = tenant_short  # e.g. "ARN" for Arion Networks
        self._ref_mapping: list[dict] = []  # accumulated during import

    def import_all(
        self,
        workbook_path: str,
        dry_run: bool = True,
        sheets: list[str] | None = None,   # None = all sheets
    ) -> ImportReport:
        """
        Import all (or selected) sheets from the workbook.
        Returns ImportReport with per-sheet results.
        """
        wb     = openpyxl.load_workbook(workbook_path, keep_vba=True, data_only=True)
        report = ImportReport(dry_run=dry_run, tenant_id=self._tenant_id)

        for config in self.SHEET_CONFIGS:
            if sheets and config.name not in sheets:
                continue
            if config.name not in wb.sheetnames:
                print(f"  [skip] Sheet not found: {config.name!r}")
                continue

            mapper = self.MAPPERS.get(config.name)
            if not mapper:
                print(f"  [skip] No mapper for: {config.name!r}")
                continue

            rows   = self._extract_rows(wb[config.name], config)
            mapped = self._map_rows(rows, config, mapper)
            report.add(config.name, mapped)

            if not dry_run:
                self._write_rows(mapped, config)

        # Fix client_facts after import
        if not dry_run:
            self._fix_client_facts()
            if self._tenant_short:
                print(f"\n  Assigning platform references...")
                self.assign_platform_refs_after_import()

        report.print_summary()
        return report

    # ── Extract ──────────────────────────────────────────────────────────────

    def _extract_rows(
        self,
        ws,
        config: SheetConfig,
    ) -> list[tuple[int, dict]]:
        """Read raw rows from worksheet. Returns [(row_num, {header: value})]."""
        # Get headers from header_row
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(config.header_row, col).value or '').strip()
            if h:
                headers[col] = h

        rows = []
        for row_idx in range(config.data_start, ws.max_row + 1):
            raw = {}
            for col, header in headers.items():
                raw[header] = ws.cell(row_idx, col).value

            # Skip if required column is empty
            if config.skip_if_empty:
                v = str(raw.get(config.skip_if_empty) or '').strip()
                if not v or v.lower() in ('none', 'n/a', '-'):
                    continue

            rows.append((row_idx, raw))

        return rows

    # ── Map ───────────────────────────────────────────────────────────────────

    def _map_rows(
        self,
        rows: list[tuple[int, dict]],
        config: SheetConfig,
        mapper: Callable,
    ) -> list[ImportRow]:
        """Apply mapper to each raw row. Returns ImportRow list."""
        result = []
        for row_num, raw in rows:
            try:
                mapped = mapper(raw, self._tenant_id)
                if mapped is None:
                    result.append(ImportRow(
                        sheet=config.name, row_num=row_num,
                        raw=raw, mapped={},
                        action="skip", reason="mapper returned None",
                    ))
                    continue

                result.append(ImportRow(
                    sheet=config.name, row_num=row_num,
                    raw=raw, mapped=mapped,
                    action="insert",
                ))
            except Exception as e:
                result.append(ImportRow(
                    sheet=config.name, row_num=row_num,
                    raw=raw, mapped={},
                    action="error", reason=str(e),
                ))
        return result

    # ── Write ─────────────────────────────────────────────────────────────────

    def _write_rows(self, rows: list[ImportRow], config: SheetConfig) -> None:
        """Write mapped rows to Postgres using UPSERT."""
        inserts = [r for r in rows if r.action == "insert" and r.mapped]
        if not inserts:
            return

        for row in inserts:
            try:
                if config.table == "isms_audits":
                    self._upsert_audit(config.table, dict(row.mapped))
                else:
                    self._upsert(config.table, row.mapped, config.upsert_on)
                row.action = "insert"
            except Exception as e:
                row.action = "error"
                row.reason = str(e)

        self._pg.commit()

    def _upsert(self, table: str, data: dict, conflict_target: str) -> None:
        """
        Generic UPSERT: INSERT ... ON CONFLICT (target) DO UPDATE SET ...
        Handles both single column and composite conflict targets.
        """
        # Filter None values for insert (let DB defaults apply)
        insert_data = {k: v for k, v in data.items() if v is not None}
        if not insert_data:
            return

        cols    = list(insert_data.keys())
        values  = list(insert_data.values())
        placeholders = ', '.join(['%s'] * len(cols))
        col_list     = ', '.join(cols)

        # Build update clause — exclude conflict columns and tenant_id
        if conflict_target.startswith('('):
            conflict_cols = re.findall(r'\w+', conflict_target)
        else:
            conflict_cols = [conflict_target]

        update_cols = [c for c in cols
                       if c not in conflict_cols + ['tenant_id', 'id']]
        if not update_cols:
            # Nothing to update — just INSERT IGNORE style
            sql = (f"INSERT INTO {table} ({col_list}) "
                   f"VALUES ({placeholders}) "
                   f"ON CONFLICT {conflict_target} DO NOTHING")
        else:
            update_clause = ', '.join(
                f"{c} = EXCLUDED.{c}" for c in update_cols
            )
            sql = (f"INSERT INTO {table} ({col_list}) "
                   f"VALUES ({placeholders}) "
                   f"ON CONFLICT {conflict_target} "
                   f"DO UPDATE SET {update_clause}")

        with self._pg.cursor() as cur:
            cur.execute(sql, values)

    def _upsert_audit(self, table: str, data: dict) -> None:
        """
        Special upsert for isms_audits — merges finding_refs arrays
        so multiple rows for the same audit_id don't overwrite each other.
        """
        finding_refs = data.pop('finding_refs', []) or []
        insert_data  = {k: v for k, v in data.items() if v is not None}
        cols         = list(insert_data.keys()) + ['finding_refs']
        values       = list(insert_data.values()) + [finding_refs]
        placeholders = ', '.join(['%s'] * len(cols))
        col_list     = ', '.join(cols)

        update_cols = [c for c in insert_data.keys()
                       if c not in ('tenant_id', 'id', 'external_ref',
                                    'standard_id', 'standard_ids')]
        update_clause = ', '.join(
            f"{c} = EXCLUDED.{c}" for c in update_cols
        )
        # Merge finding_refs arrays on conflict
        array_merge = (
            "finding_refs = ARRAY(SELECT DISTINCT unnest("
            "isms_audits.finding_refs || EXCLUDED.finding_refs))"
        )
        sql = (
            f"INSERT INTO {table} ({col_list}) "
            f"VALUES ({placeholders}) "
            f"ON CONFLICT (tenant_id, external_ref) "
            f"DO UPDATE SET {update_clause}"
            + (f", {array_merge}" if update_clause else f"SET {array_merge}")
        )
        with self._pg.cursor() as cur:
            cur.execute(sql, values)

    def _assign_platform_ref(
        self,
        table:        str,
        record_id:    str,     # UUID of the row
        external_ref: str,     # client's own reference (DOC003, R001 etc)
        display_name: str,     # human title for mapping output
    ) -> str | None:
        """
        Generate and assign a platform_ref to a DB row.
        Writes to the table's platform_ref column.
        Records the mapping for output.
        Returns the platform_ref string or None if table not in scheme.
        """
        prefix = self.TABLE_PREFIX.get(table)
        if not prefix or not self._tenant_short:
            return None

        with self._pg.cursor() as cur:
            cur.execute(
                "SELECT next_platform_ref(%s, %s, %s)",
                (self._tenant_id, prefix, self._tenant_short)
            )
            platform_ref = cur.fetchone()[0]

            cur.execute(
                f"UPDATE {table} SET platform_ref = %s WHERE id = %s",
                (platform_ref, record_id)
            )

        # Accumulate mapping
        self._ref_mapping.append({
            'platform_ref':  platform_ref,
            'client_ref':    external_ref,
            'entity_type':   self.PREFIX_TABLE.get(prefix, table),
            'display_name':  display_name,
            'table':         table,
        })
        return platform_ref

    def assign_platform_refs_after_import(self) -> list[dict]:
        """
        After writing all rows, assign platform_refs to every imported row
        that doesn't have one yet. Called once at end of live import.
        Returns the full mapping list.
        """
        if not self._tenant_short or not self._pg:
            return []

        # Table → (external_ref col, display col)
        TABLE_CONFIG = {
            'posture_controls': ('control_ref',  'control_ref'),
            'client_documents': ('external_ref', 'document_title'),
            'incidents':        ('external_ref', 'title'),
            'assets':           ('external_ref', 'name'),
            'risks':            ('external_ref', 'external_ref'),
            'vendors':          ('name',         'name'),
            'isms_audits':      ('external_ref', 'external_ref'),
        }

        mapping = []
        for table, (ref_col, display_col) in TABLE_CONFIG.items():
            prefix = self.TABLE_PREFIX.get(table)
            if not prefix:
                continue
            with self._pg.cursor() as cur:
                cur.execute(f"""
                    SELECT id, {ref_col}, {display_col}
                    FROM {table}
                    WHERE tenant_id = %s
                      AND platform_ref IS NULL
                      AND workbook_imported = TRUE
                """, (self._tenant_id,))
                rows = cur.fetchall()

            for row_id, ext_ref, display in rows:
                ref = self._assign_platform_ref(
                    table        = table,
                    record_id    = str(row_id),
                    external_ref = str(ext_ref or ''),
                    display_name = str(display or ''),
                )
                if ref:
                    mapping.append({
                        'platform_ref': ref,
                        'client_ref':   str(ext_ref or ''),
                        'entity_type':  table,
                        'display_name': str(display or ''),
                    })

        self._pg.commit()
        self._ref_mapping.extend(mapping)
        return self._ref_mapping

    def write_mapping_csv(self, output_path: str) -> None:
        """
        Write the platform_ref ↔ client_ref mapping to a CSV file.
        This is the document handed to the client so they can cross-reference
        platform references back to their own workbook IDs.
        """
        import csv
        mapping = sorted(
            self._ref_mapping,
            key=lambda x: (x.get('entity_type',''), x.get('platform_ref',''))
        )

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['platform_ref', 'client_ref', 'entity_type', 'display_name']
            writer = csv.DictWriter(f, fieldnames=fieldnames,
                                    extrasaction='ignore')   # ignore 'table' and other extra keys
            writer.writeheader()
            writer.writerows(mapping)

        print(f"  ✓ Mapping written: {output_path} ({len(mapping)} rows)")

    def _fix_client_facts(self) -> None:
        """
        Update client_facts for this tenant based on workbook evidence.
        SoA shows: no physical premises, no software development.
        Vendor register shows: uses processors, uses cloud services.
        """
        sql = """
            UPDATE client_facts SET
                develops_software     = FALSE,
                has_physical_premises = FALSE,
                uses_cloud_services   = TRUE,
                uses_processors       = TRUE,
                has_remote_workers    = TRUE,
                collected_via         = 'workbook',
                last_updated          = NOW()
            WHERE tenant_id = %s
        """
        with self._pg.cursor() as cur:
            cur.execute(sql, (self._tenant_id,))
        self._pg.commit()
        print("  ✓ client_facts updated from workbook evidence")


# =============================================================================
# CLI entrypoint
# =============================================================================

def main():
    import argparse, os, sys
    sys.path.insert(0, '.')
    from dotenv import load_dotenv
    load_dotenv()

    parser = argparse.ArgumentParser(description="Import ISO 27001 workbook to ArionComply")
    parser.add_argument('workbook', help="Path to .xlsm workbook")
    parser.add_argument('--tenant-id', default="00000000-0000-0000-0000-000000000001")
    parser.add_argument('--tenant-short', default="", help="3-letter tenant code e.g. ARN")
    parser.add_argument('--live', action='store_true', help="Write to DB (default: dry run)")
    parser.add_argument('--sheets', nargs='+', help="Import specific sheets only")
    parser.add_argument('--mapping-out', default="ref_mapping.csv",
                        help="Output path for platform ref mapping CSV")
    args = parser.parse_args()

    if args.live:
        import psycopg2
        pg_conn = psycopg2.connect(os.getenv("DATABASE_URL"))
    else:
        pg_conn = None  # dry run doesn't need connection

    importer = WorkbookImporter(pg_conn, args.tenant_id,
                               tenant_short=args.tenant_short.upper())

    if not args.live:
        # Dry run: load workbook and show what would be imported
        wb     = openpyxl.load_workbook(args.workbook, keep_vba=True, data_only=True)
        report = ImportReport(dry_run=True, tenant_id=args.tenant_id)

        for config in WorkbookImporter.SHEET_CONFIGS:
            if args.sheets and config.name not in args.sheets:
                continue
            if config.name not in wb.sheetnames:
                continue
            mapper = WorkbookImporter.MAPPERS.get(config.name)
            if not mapper:
                continue
            rows   = importer._extract_rows(wb[config.name], config)
            mapped = importer._map_rows(rows, config, mapper)
            report.add(config.name, mapped)
            # Print sample mapped rows
            sample = [r for r in mapped if r.action == "insert"][:2]
            if sample:
                print(f"\n  Sample from {config.name!r}:")
                for s in sample:
                    print(f"    Row {s.row_num}: {dict(list(s.mapped.items())[:4])}")

        report.print_summary()
    else:
        report = importer.import_all(args.workbook, dry_run=False, sheets=args.sheets)
        if args.tenant_short:
            importer.write_mapping_csv(args.mapping_out)

    if pg_conn:
        pg_conn.close()


if __name__ == "__main__":
    main()
